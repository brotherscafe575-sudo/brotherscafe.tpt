import re
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils import timezone
from django.db import transaction
from django.db.models import Sum, ExpressionWrapper, DecimalField, F as F_db, Avg as models_Avg

# -- GF(256) lookup tables for QR code generation -----------------------------
GF256_EXP = [0] * 512
GF256_LOG  = [0] * 256
_gfx = 1
for _gfi in range(255):
    GF256_EXP[_gfi] = _gfx
    GF256_LOG[_gfx] = _gfi
    _gfx <<= 1
    if _gfx & 0x100: _gfx ^= 0x11d
for _gfi in range(255, 512): GF256_EXP[_gfi] = GF256_EXP[_gfi - 255]
# -----------------------------------------------------------------------------
from .models import Table, Category, MenuItem, Order, OrderItem, ShopSettings, Discount, CustomerProfile, Combo, ComboItem, PosDraft, OfferBanner, CartOffer
import json
import openpyxl
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP


# ===================================================================
# CUSTOMER LOGIN & HISTORY
# ===================================================================

def _in_window(qs):
    """Limit a promotional queryset to whatever is running today (local date)."""
    from django.db.models import Q
    today = timezone.localdate()   # local date, not UTC
    return qs.filter(
        Q(valid_from__isnull=True) | Q(valid_from__lte=today)
    ).filter(
        Q(valid_to__isnull=True) | Q(valid_to__gte=today)
    )


def active_offer_banners():
    return _in_window(OfferBanner.objects.filter(is_active=True)).prefetch_related('menu_items')


def active_cart_offers():
    return _in_window(CartOffer.objects.filter(is_active=True))


def get_water_bottle(order_type=None):
    """The item behind the cart's quick-add card.

    Staff can now name it explicitly in Shop Settings. The old code fell through
    to "any beverage" when no water item existed, which is why the card showed
    Coffee. If nothing sensible is configured we return None and the card is
    simply hidden - showing the wrong product is worse than showing nothing.
    """
    shop = ShopSettings.objects.first()
    chosen = shop.water_bottle_item if shop else None
    if chosen:
        if order_type == 'takeaway' and not chosen.is_available_takeaway:
            return None
        if order_type == 'dine_in' and not chosen.is_available_dine_in:
            return None
        return chosen

    qs = MenuItem.objects.filter(is_water_bottle=True)
    if order_type == 'takeaway':
        qs = qs.filter(is_available_takeaway=True)
    elif order_type == 'dine_in':
        qs = qs.filter(is_available_dine_in=True)
    bottle = qs.order_by('id').first()
    if bottle:
        return bottle

    # Fallback: any item named like a water bottle (flag not yet ticked).
    qs = MenuItem.objects.filter(name__icontains='water bottle')
    if order_type == 'takeaway':
        qs = qs.filter(is_available_takeaway=True)
    elif order_type == 'dine_in':
        qs = qs.filter(is_available_dine_in=True)
    bottle = qs.order_by('id').first()
    if bottle:
        return bottle

    # No water item configured or found - the card stays hidden.
    return None


ACTIVE_ORDER_STATUSES = ['pending', 'accepted', 'preparing', 'ready']


def _get_chain(root_order):
    # All orders in the session: root + reorders ordered by time.
    return [root_order] + list(root_order.reorders.order_by('created_at'))


def _session_total(root_order):
    # Grand total of the session = sum of each order's own total.
    return sum(o.total_amount for o in _get_chain(root_order) if o.status != 'cancelled')


def _can_view_order(request, order):
    """Order IDs are sequential, so `/order/1/bill/` upward used to expose every
    customer's name, phone and items. A bill is now visible only to staff, to the
    session that placed it, or to someone holding the order's access token."""
    if request.user.is_authenticated and request.user.is_staff:
        return True
    root = order
    while root.parent_order_id:
        root = root.parent_order
    token = (request.GET.get('t') or '').strip()
    if token and token == str(root.access_token):
        return True
    session_phone = (request.session.get('customer_phone') or '').strip()
    if session_phone and session_phone in (order.customer_phone, root.customer_phone):
        return True
    return False


def _deny_order(request):
    wants_json = (request.GET.get('json')
                  or request.headers.get('x-requested-with') == 'XMLHttpRequest'
                  or 'application/json' in (request.headers.get('accept') or ''))
    if wants_json:
        return JsonResponse({'error': 'Not allowed'}, status=403)
    return render(request, 'restaurant/order_not_found.html', status=403)


def _offer_price_for(menu_item, banner):
    """The real discounted price of `menu_item` under `banner`, computed here
    rather than taken from the browser."""
    price = menu_item.price
    if not banner:
        return price
    if banner.offer_type == 'percent' and banner.off_percent:
        return max(Decimal('0'), (price - price * banner.off_percent / 100)).quantize(Decimal('1'))
    if banner.offer_type == 'flat' and banner.flat_amount:
        return max(Decimal('0'), price - banner.flat_amount).quantize(Decimal('1'))
    return price


def get_shop():
    return ShopSettings.objects.first()


def get_root_order(order):
    root = order
    while root.parent_order:
        root = root.parent_order
    return root


def get_active_order_chain(root_order):
    return [root_order] + list(root_order.reorders.filter(status__in=ACTIVE_ORDER_STATUSES).order_by('created_at'))


def get_full_order_chain(root_order):
    return [root_order] + list(root_order.reorders.order_by('created_at'))


def guest_login(request):
    """Guest login - no registration needed. Sets a guest session and redirects."""
    import random as _rand
    next_url = request.GET.get('next', '') or '/customer/start/'
    # Assign a unique guest phone so orders can be tracked within the session
    guest_phone = f"GUEST{_rand.randint(100000, 999999)}"
    request.session['customer_phone'] = guest_phone
    request.session['customer_name']  = 'Guest'
    request.session['customer_id']    = None
    request.session['is_guest']       = True
    return redirect(next_url)


def customer_login(request):
    """
    Combined Register / Login page.
    Register: name + phone → creates CustomerProfile, sets session.
    Login: phone only → looks up existing profile, sets session.
    """
    next_url = request.GET.get('next', '') or '/customer/start/'
    shop = get_shop()

    # Already logged in → go choose dine-in / takeaway
    if request.session.get('customer_phone'):
        return redirect(next_url)

    context = {'next': next_url, 'shop': shop}

    if request.method == 'POST':
        action = request.POST.get('action', 'register')
        next_url = request.POST.get('next', '') or '/customer/start/'

        if action == 'register':
            name  = request.POST.get('name', '').strip()
            phone = request.POST.get('phone', '').strip()
            context['reg_name']  = name
            context['reg_phone'] = phone

            if not name:
                context['reg_error'] = 'Please enter your name.'
            elif not phone or len(phone) < 8:
                context['reg_error'] = 'Please enter a valid phone number.'
            elif CustomerProfile.objects.filter(phone=phone).exists():
                context['reg_error'] = 'This phone number is already registered. Please use Login.'
            else:
                profile = CustomerProfile.objects.create(name=name, phone=phone)
                request.session['customer_phone'] = phone
                request.session['customer_name']  = name
                request.session['customer_id']    = profile.id
                return redirect(next_url)

        elif action == 'login':
            phone = request.POST.get('phone', '').strip()
            context['login_phone'] = phone

            if not phone or len(phone) < 8:
                context['login_error'] = 'Please enter a valid phone number.'
            else:
                try:
                    profile = CustomerProfile.objects.get(phone=phone)
                    profile.visit_count += 1
                    profile.save()
                    request.session['customer_phone'] = phone
                    request.session['customer_name']  = profile.name
                    request.session['customer_id']    = profile.id
                    return redirect(next_url)
                except CustomerProfile.DoesNotExist:
                    context['login_error'] = 'Phone number not found. Please register first.'

    return render(request, 'restaurant/customer_login.html', context)


def order_type_select(request):
    """After login / guest: ask Dine In (pick a table) or Takeaway."""
    if not request.session.get('customer_phone'):
        return redirect('/customer/login/?next=/customer/start/')
    # Dine In goes straight to the scanned table's menu (QR = table).
    # If no QR was scanned, default to the first table (Table 1) - never ask.
    dine_table = None
    sid = request.session.get('scanned_table_id')
    if sid:
        dine_table = Table.objects.filter(id=sid, is_active=True).first()
    if not dine_table:
        dine_table = Table.objects.filter(is_active=True).order_by('number').first()
    return render(request, 'restaurant/order_type.html', {
        'shop': get_shop(),
        'dine_table': dine_table,
        'name': request.session.get('customer_name', ''),
        'is_guest': request.session.get('is_guest', False),
    })


def customer_logout(request):
    request.session.flush()
    return redirect('/')


def customer_history(request):
    phone = request.session.get('customer_phone')
    if not phone:
        return redirect("/customer/login/?next=/customer/history/")
    shop = get_shop()
    name = request.session.get('customer_name', '')

    all_orders = Order.objects.filter(
        customer_phone=phone,
        parent_order__isnull=True,
    ).prefetch_related('items__menu_item','reorders__items__menu_item').order_by('-created_at')

    def enrich(order):
        linked = [order] + list(order.reorders.order_by('created_at'))
        # After a reorder is marked Ready, its items are MERGED into the root order
        # and the child is set to status='completed'. So root.total_amount already
        # reflects ALL items. Summing child totals on top would double-count.
        # Only include child totals for reorders that are still active (not yet merged).
        ACTIVE = {'pending', 'accepted', 'preparing', 'ready'}
        active_children = [o for o in linked[1:] if o.status in ACTIVE]
        combined_total = order.total_amount + sum(o.total_amount for o in active_children)
        all_items = []
        for o in linked:
            for item in o.items.select_related('menu_item').all():
                item.order = o
                all_items.append(item)
        return {
            'order': order,
            'combined_total': combined_total,
            'all_items': all_items,
            'all_orders': linked,      # for live card to show per-order grouping
            'reorder_count': len(linked)-1,
        }

    # Active = any non-completed/non-cancelled root order
    ACTIVE_STATUSES = ['pending','accepted','preparing','ready']
    active_entry = None
    past_entries = []
    for o in all_orders:
        e = enrich(o)
        if o.status in ACTIVE_STATUSES and active_entry is None:
            active_entry = e
        else:
            past_entries.append(e)

    all_entries = [enrich(o) for o in all_orders]

    # Stats - only root orders to avoid double-counting reorder child orders
    completed_count = sum(1 for o in all_orders if o.status == 'completed')
    from django.db.models import Sum
    total_spent = Order.objects.filter(
        customer_phone=phone,
        status='completed',
        parent_order__isnull=True,
    ).aggregate(t=Sum('total_amount'))['t'] or 0

    # Last table - prefer session value set when menu was loaded (for refresh support)
    last_order = all_orders.first()
    sess_table = request.session.get('last_table_id')
    sess_type  = request.session.get('last_order_type', '')
    if sess_table:
        last_table_id = sess_table
        is_takeaway   = False
    elif sess_type == 'takeaway':
        last_table_id = None
        is_takeaway   = True
    else:
        last_table_id = last_order.table.id if last_order and last_order.table else None
        is_takeaway   = last_order.order_type == 'takeaway' if last_order else False

    prog_steps = [
        ('pending',   '📋', 'Received'),
        ('accepted',  '✅', 'Accepted'),
        ('preparing', '👨‍🍳', 'Cooking'),
        ('ready',     '🔔', 'Ready'),
        ('completed', '🎉', 'Done'),
    ]

    return render(request, 'restaurant/customer_home.html', {
        'orders': all_entries,
        'active_order': active_entry,
        'past_orders': past_entries,
        'customer_name': name,
        'customer_phone': phone,
        'completed_count': completed_count,
        'total_spent': total_spent,
        'last_table_id': last_table_id,
        'is_takeaway': is_takeaway,
        'prog_steps': prog_steps,
        'shop': shop,
    })


def table_selection(request):
    """Root URL - redirect customers to login, show nothing useful."""
    # If already logged in, show a simple welcome with history link
    if request.session.get('customer_phone'):
        return redirect('/customer/start/')
    return redirect('/customer/login/?next=/customer/start/')


def _get_active_order_for_menu(phone):
    """Return active order entry for a customer phone, or None."""
    if not phone:
        return None
    ACTIVE_STATUSES = ['pending', 'accepted', 'preparing', 'ready']
    order = Order.objects.filter(
        customer_phone=phone,
        parent_order__isnull=True,
        status__in=ACTIVE_STATUSES,
    ).prefetch_related('items__menu_item', 'reorders__items__menu_item').order_by('-created_at').first()
    if not order:
        return None
    linked = [order] + list(order.reorders.order_by('created_at'))
    combined_subtotal = sum(o.subtotal for o in linked)
    combined_parcel   = sum(o.parcel_charge for o in linked)
    combined_discount = sum(o.discount_amount for o in linked)
    combined_total    = combined_subtotal + combined_parcel - combined_discount
    return {'order': order, 'combined_total': combined_total, 'combined_parcel': combined_parcel, 'all_orders': linked}


def _get_order_history_for_menu(phone, limit=5):
    """Return past orders for a customer (completed/cancelled), newest first."""
    if not phone:
        return []
    past = Order.objects.filter(
        customer_phone=phone,
        parent_order__isnull=True,
        status__in=['completed', 'cancelled'],
    ).prefetch_related('items__menu_item', 'reorders__items__menu_item').order_by('-created_at')[:limit]
    result = []
    for o in past:
        linked = [o] + list(o.reorders.order_by('created_at'))
        combined_total = sum(x.subtotal + x.parcel_charge - x.discount_amount for x in linked)
        all_items = []
        for x in linked:
            for item in x.items.select_related('menu_item').all():
                all_items.append(item)
        result.append({'order': o, 'combined_total': combined_total, 'all_items': all_items, 'reorder_count': len(linked) - 1})
    return result


@ensure_csrf_cookie
def menu_view(request, table_id):
    # Require customer login (skip for reorders - session already set)
    reorder_from = request.GET.get('reorder_from')
    if not request.session.get('customer_phone') and not reorder_from:
        # Remember which table QR was scanned so "Dine In" on the choice screen
        # jumps straight back to this table - but still ask Dine In vs Takeaway
        # first, since a customer may want takeaway even after scanning a table QR.
        request.session['scanned_table_id'] = table_id
        return redirect("/customer/login/?next=/customer/start/")

    table = get_object_or_404(Table, id=table_id, is_active=True)
    # Remember the scanned table for this session (QR = table)
    request.session['scanned_table_id'] = table.id
    categories = Category.objects.filter(is_active=True).prefetch_related('items')
    discounts = Discount.objects.filter(is_active=True)
    shop = get_shop()
    reorder_order = None
    reorder_customer_name = ''
    reorder_customer_phone = ''
    cart = '{}'
    cart_count = 0
    if reorder_from:
        try:
            # Walk up to root order so we always link to the original
            reorder_order = Order.objects.get(id=reorder_from)
            while reorder_order.parent_order:
                reorder_order = reorder_order.parent_order
            if not _can_view_order(request, reorder_order):
                return _deny_order(request)
            reorder_customer_name = reorder_order.customer_name
            reorder_customer_phone = reorder_order.customer_phone
            prefill = {
                str(item.menu_item.id): {
                    'name': item.menu_item.name,
                    'price': float(item.unit_price),
                    'qty': item.quantity,
                }
                for item in reorder_order.items.select_related('menu_item').all()
            }
            if prefill:
                cart = json.dumps(prefill)
                cart_count = sum(v['qty'] for v in prefill.values())
        except Order.DoesNotExist:
            reorder_order = None

    phone = request.session.get('customer_phone', '')
    active_order = _get_active_order_for_menu(phone)
    past_orders = _get_order_history_for_menu(phone)
    # Pre-annotate categories with non-water-bottle item count for the template
    _categories_qs = Category.objects.filter(is_active=True).prefetch_related(
        'items').order_by('order', 'name')
    _is_takeaway = locals().get('is_takeaway', False)
    for _cat in _categories_qs:
        _cat.non_wbottle_count = sum(
            1 for i in _cat.items.all()
            if not i.is_water_bottle and (
                (_is_takeaway and i.is_available_takeaway) or
                (not _is_takeaway and i.is_available_dine_in)
            )
        )
    categories = _categories_qs

    _all_combos = list(Combo.objects.filter(is_active=True).prefetch_related('combo_items__menu_item'))
    # Combo Offers get their own highlighted strip at the top of the menu, so
    # keep them out of the ordinary combos carousel further down the page.
    combo_offers = [c for c in _all_combos if c.is_live_offer]
    combos = [c for c in _all_combos if not c.is_offer]

    # Store last menu URL in session so page refresh works
    request.session['last_table_id'] = table_id
    request.session['last_order_type'] = 'dine_in'

    wb = get_water_bottle(order_type='dine_in')
    wb_enabled = (shop.show_water_bottle_in_cart if shop else True) and wb is not None
    wb_price = float(shop.water_bottle_cart_price) if shop and shop.water_bottle_cart_price else (float(wb.price) if wb else 10)
    # Build offer banners JSON for the inline panel
    offer_banners_qs = active_offer_banners()
    offer_banners_json = json.dumps([{
        'id': b.id, 'title': b.title, 'subtitle': b.subtitle,
        'offer_type': b.offer_type, 'off_percent': float(b.off_percent),
        'flat_amount': float(b.flat_amount), 'offer_label': b.offer_label,
        'emoji': b.emoji, 'bg_color': b.bg_color, 'image_url': b.banner_image_url,
        'items': [{'id': mi.id, 'name': mi.name, 'price': float(mi.price),
                   'category': mi.category.name,
                   'image': mi.image.url if mi.image else ''} for mi in b.menu_items.filter(is_available_dine_in=True)]
    } for b in offer_banners_qs])
    # Cart offers
    cart_offers_json = json.dumps([{
        'id': co.id, 'title': co.title, 'subtitle': co.subtitle,
        'min_cart_value': float(co.min_cart_value), 'reward_type': co.reward_type,
        'percent_off': float(co.percent_off), 'flat_off': float(co.flat_off),
        'free_item_id': co.free_item_id, 'free_item_name': co.free_item.name if co.free_item else '',
        'free_item_price': float(co.free_item.price) if co.free_item else 0,
        'emoji': co.emoji, 'reward_label': co.reward_label,
    } for co in active_cart_offers()])
    return render(request, 'restaurant/menu.html', {
        'offer_banners': offer_banners_qs,
        'offer_banners_json': offer_banners_json,
        'cart_offers_json': cart_offers_json,
        'water_bottle_enabled': wb_enabled,
        'water_bottle_price': wb_price,
        'table': table, 'categories': categories,
        'discounts': discounts, 'shop': shop,
        'reorder_from': str(reorder_order.id) if reorder_order else None,
        'reorder_order': reorder_order,
        'cart': cart,
        'cart_count': cart_count,
        'is_takeaway': False,
        'active_order': active_order,
        'past_orders': past_orders,
        'combos': combos,
        'combo_offers': combo_offers,
        'water_bottle_id': wb.id if wb else None,
        'water_bottle_name': wb.name if wb else 'Water Bottle',
    })


@ensure_csrf_cookie
def takeaway_menu(request):
    """Takeaway order - no table needed"""
    if not request.session.get('customer_phone'):
        return redirect("/customer/login/?next=/takeaway/")
    categories = Category.objects.filter(is_active=True).prefetch_related('items')
    discounts = Discount.objects.filter(is_active=True)
    shop = get_shop()
    phone = request.session.get('customer_phone', '')
    active_order = _get_active_order_for_menu(phone)
    past_orders = _get_order_history_for_menu(phone)
    _all_combos = list(Combo.objects.filter(is_active=True).prefetch_related('combo_items__menu_item'))
    # Combo Offers get their own highlighted strip at the top of the menu, so
    # keep them out of the ordinary combos carousel further down the page.
    combo_offers = [c for c in _all_combos if c.is_live_offer]
    combos = [c for c in _all_combos if not c.is_offer]
    request.session['last_table_id'] = None
    request.session['last_order_type'] = 'takeaway'
    wb = get_water_bottle(order_type='takeaway')
    # Build parcel charges map for frontend: {item_id: parcel_charge}
    default_pc = shop.default_parcel_charge if shop else 0
    parcel_map = {}
    for cat in categories:
        for item in cat.items.all():
            pc = float(item.parcel_charge) if item.parcel_charge else float(default_pc)
            parcel_map[item.id] = pc  # include all items, 0 means no charge

    import json as _json
    offer_banners_qs2 = active_offer_banners()
    offer_banners_json2 = _json.dumps([{
        'id': b.id, 'title': b.title, 'subtitle': b.subtitle,
        'offer_type': b.offer_type, 'off_percent': float(b.off_percent),
        'flat_amount': float(b.flat_amount), 'offer_label': b.offer_label,
        'emoji': b.emoji, 'bg_color': b.bg_color, 'image_url': b.banner_image_url,
        'items': [{'id': mi.id, 'name': mi.name, 'price': float(mi.price),
                   'category': mi.category.name,
                   'image': mi.image.url if mi.image else ''} for mi in b.menu_items.filter(is_available_takeaway=True)]
    } for b in offer_banners_qs2])
    cart_offers_json2 = _json.dumps([{
        'id': co.id, 'title': co.title, 'subtitle': co.subtitle,
        'min_cart_value': float(co.min_cart_value), 'reward_type': co.reward_type,
        'percent_off': float(co.percent_off), 'flat_off': float(co.flat_off),
        'free_item_id': co.free_item_id, 'free_item_name': co.free_item.name if co.free_item else '',
        'free_item_price': float(co.free_item.price) if co.free_item else 0,
        'emoji': co.emoji, 'reward_label': co.reward_label,
    } for co in active_cart_offers()])
    return render(request, 'restaurant/menu.html', {
        'offer_banners': offer_banners_qs2,
        'offer_banners_json': offer_banners_json2,
        'cart_offers_json': cart_offers_json2,
        'table': None, 'categories': categories,
        'discounts': discounts, 'shop': shop,
        'cart': '{}', 'cart_count': 0,
        'is_takeaway': True,
        'active_order': active_order,
        'past_orders': past_orders,
        'combos': combos,
        'combo_offers': combo_offers,
        'water_bottle_id': wb.id if wb else None,
        'water_bottle_name': wb.name if wb else 'Water Bottle',
        'water_bottle_enabled': (shop.show_water_bottle_in_cart if shop else True) and wb is not None,
        'water_bottle_price': (float(shop.water_bottle_cart_price)
                               if shop and shop.water_bottle_cart_price
                               else (float(wb.price) if wb else 0)),
        'default_parcel_charge': float(default_pc),
        'parcel_map_json': _json.dumps(parcel_map),
    })


def place_order(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body)
        customer_name  = (data.get('customer_name') or request.session.get('customer_name', '')).strip()
        customer_phone = (data.get('customer_phone') or request.session.get('customer_phone', '')).strip()
        special_instructions = data.get('special_instructions', '')
        items = data.get('items', [])
        discount_percent = Decimal(str(data.get('discount_percent', 0)))
        cart_offer_flat = Decimal(str(data.get('cart_offer_flat', 0) or 0))
        cart_offer_title = data.get('cart_offer_title', '')
        cart_offer_id = data.get('cart_offer_id')
        order_type = data.get('order_type', 'dine_in')
        table_id = data.get('table_id')
        parent_order_id = data.get('parent_order_id')
        payment_method = data.get('payment_method')
        cash_received = data.get('cash_received')
        change_amount = data.get('change_amount')

        if not customer_name:
            return JsonResponse({'success': False, 'error': 'Customer name is required'})
        if not customer_phone:
            return JsonResponse({'success': False, 'error': 'Phone number is required'})
        if not items:
            return JsonResponse({'success': False, 'error': 'No items in order'})

        table = None
        if order_type == 'dine_in' and table_id:
            table = Table.objects.filter(id=table_id, is_active=True).first()
            if not table:
                return JsonResponse({'success': False, 'error': 'Table not found'})

        special_note = special_instructions
        parent_order_obj = None
        if parent_order_id:
            try:
                parent_order_obj = Order.objects.get(id=parent_order_id)
                # Always link to root order
                while parent_order_obj.parent_order_id:
                    parent_order_obj = parent_order_obj.parent_order
                special_note = f'Reorder (linked to #{parent_order_obj.order_number})' + (f' - {special_instructions}' if special_instructions else '')
            except Order.DoesNotExist:
                pass

        import random
        order = Order(
            table=table,
            customer_name=customer_name,
            customer_phone=customer_phone,
            special_instructions=special_note,
            discount_percent=discount_percent,
            order_type=order_type,
            order_number=f"BC{timezone.now().strftime('%Y%m%d')}{random.randint(1000,9999)}",
            subtotal=Decimal('0'),
            discount_amount=Decimal('0'),
            total_amount=Decimal('0'),
            parent_order=parent_order_obj,
        )
        # Only a logged-in staff member (POS) may declare an order already paid.
        # Previously any customer could post payment_method:'offline' and get a
        # completed, paid order without paying anything.
        is_staff_request = request.user.is_authenticated and request.user.is_staff
        if is_staff_request and payment_method in ['offline', 'online']:
            order.payment_method = payment_method
            order.payment_status = 'paid_offline' if payment_method == 'offline' else 'paid_online'
            order.status = 'completed'
            order.completed_at = timezone.now()
        if cash_received is not None:
            order.cash_received = Decimal(str(cash_received))
        if change_amount is not None:
            order.change_amount = Decimal(str(change_amount))
        order.save()

        subtotal = Decimal('0')

        # Batch-fetch all needed menu items and combos in 2 queries (not N queries)
        regular_ids = [item_data['id'] for item_data in items
                       if not item_data.get('combo_id') and item_data.get('id')]
        combo_ids   = [item_data['combo_id'] for item_data in items if item_data.get('combo_id')]

        # Filter items by availability based on order_type
        if order_type == 'takeaway':
            menu_items_map = {m.id: m for m in MenuItem.objects.filter(id__in=regular_ids, is_available_takeaway=True)} if regular_ids else {}
            fallback_item  = MenuItem.objects.filter(is_available_takeaway=True).first() if combo_ids else None
        else:  # dine_in
            menu_items_map = {m.id: m for m in MenuItem.objects.filter(id__in=regular_ids, is_available_dine_in=True)} if regular_ids else {}
            fallback_item  = MenuItem.objects.filter(is_available_dine_in=True).first() if combo_ids else None
        
        combos_map     = {c.id: c for c in Combo.objects.filter(id__in=combo_ids, is_active=True).prefetch_related('combo_items__menu_item')} if combo_ids else {}

        # -- Offer context, resolved server-side ----------------------------
        # Which single item (if any) this order is allowed to receive for free,
        # and what each banner-discounted item is really allowed to cost.
        free_item_grant_id = None
        active_cart_offer = None
        if cart_offer_id:
            active_cart_offer = active_cart_offers().filter(id=cart_offer_id).first()
            if active_cart_offer and active_cart_offer.reward_type == 'free_item':
                free_item_grant_id = active_cart_offer.free_item_id

        offer_banner_map = {}
        for _b in active_offer_banners():
            for _mi in _b.menu_items.all():
                offer_banner_map.setdefault(_mi.id, _b)

        # A percentage the browser asked for is only honoured if a matching
        # active Discount (or the shop default) actually exists. Otherwise any
        # customer could post discount_percent: 90 and pay a tenth of the bill.
        if discount_percent > 0 and not cart_offer_id:
            shop_default = (get_shop().default_discount_percent or Decimal('0')) if get_shop() else Decimal('0')
            today_ = timezone.localdate()   # IST local date
            allowed_pcts = {
                d.percent for d in Discount.objects.filter(is_active=True)
                if (not d.valid_from or d.valid_from <= today_)
                and (not d.valid_to or d.valid_to >= today_)
            }
            allowed_pcts.add(shop_default)
            if discount_percent not in allowed_pcts:
                discount_percent = Decimal('0')

        order_items_to_create = []
        for item_data in items:
            combo_id = item_data.get('combo_id')
            if combo_id:
                combo = combos_map.get(int(combo_id))
                if not combo:
                    continue
                qty = max(1, int(item_data.get('qty', 1)))
                combo_price = combo.price * qty
                combo_items = list(combo.combo_items.all())
                rep_item = combo_items[0].menu_item if combo_items else fallback_item
                if rep_item:
                    # `combo` is what every bill renders; rep_item only exists so
                    # sales reports still have a MenuItem to group by.
                    order_items_to_create.append(OrderItem(
                        order=order, menu_item=rep_item, combo=combo, quantity=qty,
                        unit_price=combo.price,
                        notes=', '.join(f'{ci.quantity}x {ci.menu_item.name}' for ci in combo_items)
                    ))
                subtotal += combo_price
                continue
            item_id = item_data.get('id')
            if not item_id:
                continue
            menu_item = menu_items_map.get(int(item_id))
            if not menu_item:
                continue
            qty = max(1, int(item_data.get('qty', 1)))
            unit_price = menu_item.price
            is_free_line = False

            # Free item granted by a cart offer. The client cannot simply declare
            # something free - we re-check it against the offer it claims to
            # come from, and that the cart actually meets the minimum.
            if item_data.get('free_offer') and free_item_grant_id == menu_item.id:
                unit_price = Decimal('0')
                qty = 1
                is_free_line = True
            elif item_data.get('offer_label') and item_data.get('unit_price') is not None:
                # Discounted price from an offer banner - recompute it server-side
                # instead of trusting whatever the browser sent.
                allowed = _offer_price_for(menu_item, offer_banner_map.get(menu_item.id))
                try:
                    provided = Decimal(str(item_data['unit_price']))
                except Exception:
                    provided = allowed
                # Accept the client price only if it is no cheaper than the real
                # offer price, so a crafted request can't zero out the cart.
                unit_price = provided if allowed <= provided <= menu_item.price else allowed
            # Store offer details in notes if present
            offer_note = ''
            offer_label = item_data.get('offer_label', '')
            offer_type = item_data.get('offer_type', '')
            offer_title = item_data.get('offer_title', '')
            if offer_label:
                offer_note = f'🏷️ {offer_label}'
                if offer_title:
                    offer_note += f' - {offer_title}'
            order_items_to_create.append(OrderItem(
                order=order, menu_item=menu_item, quantity=qty, unit_price=unit_price,
                is_free=is_free_line,
                notes='🎁 FREE with offer' if is_free_line else offer_note
            ))
            subtotal += unit_price * qty

        # Single bulk insert instead of N inserts
        OrderItem.objects.bulk_create(order_items_to_create)

        # Add parcel charges for takeaway orders
        parcel_total = Decimal('0')
        if order_type == 'takeaway':
            shop_obj = get_shop()
            default_pc = shop_obj.default_parcel_charge if shop_obj else Decimal('0')
            for oi in order_items_to_create:
                item_parcel = oi.menu_item.parcel_charge if oi.menu_item.parcel_charge else default_pc
                parcel_total += item_parcel * oi.quantity

        # subtotal = items only (no parcel), parcel stored separately
        order.subtotal = subtotal
        order.parcel_charge = parcel_total

        # Re-validate the cart offer server-side - a client may send a stale
        # cart_offer_flat/discount_percent for an offer whose min_cart_value
        # is no longer met (e.g. items removed after the offer was applied).
        if cart_offer_id:
            cart_offer_obj = active_cart_offer
            # `subtotal` here excludes the free line (it costs 0), so compare the
            # paid subtotal against the minimum - that is what unlocked the offer.
            if not cart_offer_obj or subtotal < cart_offer_obj.min_cart_value:
                cart_offer_flat = Decimal('0')
                if cart_offer_obj and cart_offer_obj.reward_type == 'percent':
                    discount_percent = Decimal('0')
                cart_offer_title = ''
                # The free item was never actually earned - drop it and its note.
                if free_item_grant_id:
                    OrderItem.objects.filter(order=order, is_free=True).delete()
            elif cart_offer_obj.reward_type == 'percent':
                # Never let the client pick its own percentage.
                discount_percent = cart_offer_obj.percent_off
            elif cart_offer_obj.reward_type == 'flat':
                cart_offer_flat = cart_offer_obj.flat_off

        order.discount_is_flat = False
        order.offer_title = cart_offer_title or ''
        # The percentage actually used must be written back onto the order.
        # Otherwise the order carries a discount *amount* with no percentage
        # behind it, and the next calculate_totals() resets it to zero - which
        # is what made a customer's offer vanish the moment staff billed it.
        order.discount_percent = discount_percent if discount_percent > 0 else Decimal('0')
        if subtotal > 0 and cart_offer_flat > 0:
            order.discount_amount = min(cart_offer_flat, subtotal).quantize(Decimal('0.01'))
            order.discount_is_flat = True
            order.discount_percent = Decimal('0')
        elif subtotal > 0 and discount_percent > 0:
            order.discount_amount = (subtotal * discount_percent / 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            order.discount_amount = Decimal('0.00')
        # Append cart offer name to special instructions for staff/bill
        if cart_offer_title and order.discount_amount > 0:
            existing = order.special_instructions or ''
            offer_note = f'🎁 Offer: {cart_offer_title}'
            order.special_instructions = (offer_note + (' | ' + existing if existing else '')) if existing != offer_note else existing
        order.total_amount = order.subtotal + order.parcel_charge - order.discount_amount
        update_kwargs = {
            'subtotal': order.subtotal,
            'parcel_charge': order.parcel_charge,
            'discount_amount': order.discount_amount,
            'discount_percent': order.discount_percent,
            'discount_is_flat': order.discount_is_flat,
            'offer_title': order.offer_title,
            'total_amount': order.total_amount,
            'special_instructions': order.special_instructions,
        }
        if order.payment_method:
            update_kwargs.update({
                'payment_method': order.payment_method,
                'payment_status': order.payment_status,
                'status': order.status,
                'completed_at': order.completed_at,
            })
        if order.cash_received is not None:
            update_kwargs['cash_received'] = order.cash_received
        if order.change_amount is not None:
            update_kwargs['change_amount'] = order.change_amount

        Order.objects.filter(pk=order.pk).update(**update_kwargs)
        if table:
            Table.objects.filter(pk=table.pk).update(status='occupied')

        return JsonResponse({'success': True, 'order_id': order.id, 'order_number': order.order_number, 'history_url': '/customer/history/'})

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid request data'}, status=400)
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@ensure_csrf_cookie
def reorder_menu(request, order_id):
    """Show menu so customer can pick new items - these will be added to existing order bill"""
    order = get_object_or_404(Order, id=order_id)
    if not _can_view_order(request, order):
        return _deny_order(request)
    categories = Category.objects.filter(is_active=True).prefetch_related('items')
    discounts = Discount.objects.filter(is_active=True)
    shop = get_shop()
    import json as _json
    prefill = {str(item.menu_item.id): {'name': item.menu_item.name, 'price': float(item.unit_price), 'qty': item.quantity}
               for item in order.items.all()}
    offer_banners_qs3 = active_offer_banners()
    offer_banners_json3 = _json.dumps([{
        'id': b.id, 'title': b.title, 'subtitle': b.subtitle,
        'offer_type': b.offer_type, 'off_percent': float(b.off_percent),
        'flat_amount': float(b.flat_amount), 'offer_label': b.offer_label,
        'emoji': b.emoji, 'bg_color': b.bg_color, 'image_url': b.banner_image_url,
        'items': [{'id': mi.id, 'name': mi.name, 'price': float(mi.price),
                   'category': mi.category.name,
                   'image': mi.image.url if mi.image else ''} for mi in b.menu_items.filter(is_available_dine_in=True)]
    } for b in offer_banners_qs3])
    cart_offers_json3 = _json.dumps([{
        'id': co.id, 'title': co.title, 'subtitle': co.subtitle,
        'min_cart_value': float(co.min_cart_value), 'reward_type': co.reward_type,
        'percent_off': float(co.percent_off), 'flat_off': float(co.flat_off),
        'free_item_id': co.free_item_id, 'free_item_name': co.free_item.name if co.free_item else '',
        'free_item_price': float(co.free_item.price) if co.free_item else 0,
        'emoji': co.emoji, 'reward_label': co.reward_label,
    } for co in active_cart_offers()])
    return render(request, 'restaurant/menu.html', {
        'offer_banners': offer_banners_qs3,
        'offer_banners_json': offer_banners_json3,
        'cart_offers_json': cart_offers_json3,
        'table': order.table,
        'categories': categories,
        'discounts': discounts,
        'shop': shop,
        'cart': _json.dumps(prefill),
        'cart_count': sum(v['qty'] for v in prefill.values()),
        'reorder_for': order,
        'is_reorder': True,
    })


def add_items_to_order(request, order_id):
    """Add new items from reorder menu into the existing order bill"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        order = get_object_or_404(Order, id=order_id)
        data = json.loads(request.body)
        items = data.get('items', [])
        order_type = data.get('order_type', order.order_type)  # Use order's type if not specified
        if not items:
            return JsonResponse({'success': False, 'error': 'No items provided'})
        for item_data in items:
            combo_id = item_data.get('combo_id')
            if combo_id:
                try:
                    combo = Combo.objects.prefetch_related('combo_items__menu_item').get(id=combo_id, is_active=True)
                except Combo.DoesNotExist:
                    continue
                qty = max(1, int(item_data.get('qty', 1)))
                combo_price = combo.price * qty
                combo_items = list(combo.combo_items.select_related('menu_item').all())
                if combo_items:
                    raw_total = sum(ci.menu_item.price * ci.quantity for ci in combo_items)
                    for ci in combo_items:
                        if raw_total > 0:
                            proportion = (ci.menu_item.price * ci.quantity) / raw_total
                            unit = (combo_price * proportion / ci.quantity).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        else:
                            unit = (combo_price / len(combo_items) / ci.quantity).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        OrderItem.objects.create(
                            order=order, menu_item=ci.menu_item,
                            quantity=ci.quantity * qty, unit_price=unit,
                            notes=f'[Combo: {combo.name}]'
                        )
                continue
            menu_item = MenuItem.objects.filter(id=item_data['id']).first()
            if not menu_item:
                continue
            # Check availability based on order type
            if order_type == 'takeaway' and not menu_item.is_available_takeaway:
                continue
            if order_type == 'dine_in' and not menu_item.is_available_dine_in:
                continue
            qty = max(1, int(item_data.get('qty', 1)))
            existing = order.items.filter(menu_item=menu_item).first()
            if existing:
                existing.quantity += qty
                existing.save()
            else:
                OrderItem.objects.create(order=order, menu_item=menu_item,
                                         quantity=qty, unit_price=menu_item.price)
        order.calculate_totals()
        Order.objects.filter(pk=order.pk).update(
            subtotal=order.subtotal,
            discount_amount=order.discount_amount,
            total_amount=order.total_amount,
        )
        return JsonResponse({'success': True, 'order_id': order.id, 'order_number': order.order_number, 'history_url': '/customer/history/'})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def reorder(request, order_id):
    """Redirect customer to menu with original items pre-loaded for reorder"""
    orig = get_object_or_404(Order, id=order_id)
    if not _can_view_order(request, orig):
        return _deny_order(request)
    if orig.table:
        return redirect(f"/table/{orig.table.id}/menu/?reorder_from={orig.id}")
    else:
        return redirect(f"/takeaway/?reorder_from={orig.id}")


def order_status(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if not _can_view_order(request, order):
        return _deny_order(request)
    shop = get_shop()

    # Find root order and gather all orders in this session (first + reorders)
    root_order = get_root_order(order)
    all_orders = get_full_order_chain(root_order)
    active_orders = get_active_order_chain(root_order)

    # Build combined items list from active orders only.
    # Completed children have already been merged into root.items - excluding them
    # prevents double-counting. Only show root + truly active (non-completed) children.
    active_children = [o for o in active_orders if o.parent_order_id and o.status != 'completed']
    display_orders = [root_order] + active_children
    combined_items = []
    combined_subtotal = Decimal('0')
    for o in display_orders:
        for item in o.items.select_related('menu_item', 'combo').prefetch_related('combo__combo_items__menu_item').all():
            combined_items.append({
                'name': item.display_name,
                'is_combo': item.is_combo,
                'is_free': item.is_free,
                'combo_items': item.combo_components,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'total_price': item.total_price,
                'is_reorder': o.parent_order is not None,
            })
            combined_subtotal += item.total_price
    discount_percent = root_order.discount_percent
    discount_amount = (combined_subtotal * discount_percent / 100).quantize(Decimal('0.01')) if discount_percent else Decimal('0')
    combined_total = combined_subtotal - discount_amount

    # Bill is ready only when every order in session is completed
    all_completed = all(o.status == 'completed' for o in all_orders)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'status': order.status,           # this order's status (for progress bar)
            'root_status': root_order.status, # root order status
            'all_completed': all_completed,   # True only when EVERY order is done
            'payment_status': order.payment_status,
            'status_display': order.get_status_display(),
            'combined_total': float(combined_total),
            'has_reorders': len(all_orders) > 1,
        })
    return render(request, 'restaurant/order_status.html', {
        'order': order,
        'root_order': root_order,
        'all_orders': all_orders,
        'combined_items': combined_items,
        'combined_subtotal': combined_subtotal,
        'discount_percent': discount_percent,
        'discount_amount': discount_amount,
        'combined_total': combined_total,
        'has_reorders': len(all_orders) > 1,
        'all_completed': all_completed,
        'shop': shop,
    })


def bill_view(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if not _can_view_order(request, order):
        return _deny_order(request)
    shop = get_shop()
    # Find root order - bill is only available when root order is completed
    root_order = order
    while root_order.parent_order:
        root_order = root_order.parent_order
    all_orders_check = get_full_order_chain(root_order)
    # Cancelled reorders must not block the bill — only require that every
    # non-cancelled order is completed.
    non_cancelled = [o for o in all_orders_check if o.status != 'cancelled']
    all_done = all(o.status == 'completed' for o in non_cancelled)
    if not all_done:
        return render(request, 'restaurant/waiting_for_bill.html', {'order': root_order, 'shop': shop})
    all_orders = get_full_order_chain(root_order)
    active_orders = [o for o in all_orders if o.status != 'cancelled']

    # Each order keeps its own items and discount. Build per-order sections
    # so the bill shows a clear breakdown: original order + each reorder.
    order_sections = []
    grand_total = Decimal('0')
    for o in active_orders:
        items_qs = list(o.items.select_related('menu_item', 'combo').prefetch_related(
            'combo__combo_items__menu_item').all())
        sub = sum((i.total_price for i in items_qs), Decimal('0'))
        if o.discount_is_flat:
            disc = min(o.discount_amount or Decimal('0'), sub)
        elif o.discount_percent:
            disc = (sub * o.discount_percent / 100).quantize(Decimal('0.01'))
        else:
            disc = Decimal('0')
        parcel = o.parcel_charge or Decimal('0')
        sec_total = max(Decimal('0'), sub + parcel - disc)
        grand_total += sec_total
        order_sections.append({
            'order_number': o.order_number,
            'is_reorder': o.parent_order_id is not None,
            'items': [{
                'name': item.display_name,
                'is_combo': item.is_combo,
                'is_free': item.is_free,
                'combo_items': item.combo_components,
                'notes': item.notes,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'total_price': item.total_price,
                'is_reorder': o.parent_order_id is not None,
                'order_number': o.order_number,
            } for item in items_qs],
            'subtotal': sub,
            'parcel': parcel,
            'discount_amount': disc,
            'discount_percent': o.discount_percent,
            'discount_is_flat': o.discount_is_flat,
            'offer_title': o.offer_title or '',
            'section_total': sec_total,
        })

    # Flat combined_items list for templates that don't show per-section layout
    combined_items = [item for s in order_sections for item in s['items']]
    combined_subtotal = sum(s['subtotal'] for s in order_sections)
    combined_parcel   = sum(s['parcel']   for s in order_sections)
    discount_amount   = sum(s['discount_amount'] for s in order_sections)
    combined_total    = grand_total

    return render(request, 'restaurant/bill.html', {
        'order': order,
        'root_order': root_order,
        'all_orders': all_orders,
        'order_sections': order_sections,
        'combined_items': combined_items,
        'combined_subtotal': combined_subtotal,
        'combined_parcel': combined_parcel,
        'discount_amount': discount_amount,
        'combined_total': combined_total,
        'offer_title': root_order.offer_title or '',
        'discount_is_flat': root_order.discount_is_flat,
        'discount_percent': root_order.discount_percent,
        'has_reorders': len(active_orders) > 1,
        'shop': shop,
    })


def pay_online(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if not _can_view_order(request, order):
        return _deny_order(request)
    shop = get_shop()
    if order.payment_status not in ('paid_online', 'paid_offline'):
        order.payment_method = 'online'
        order.payment_status = 'online_pending'
        order.save()
    return render(request, 'restaurant/pay_online.html', {'order': order, 'shop': shop})


def confirm_payment(request, order_id):
    """Customer says they have paid via UPI. This only records an intent -
    staff still confirm it from the portal. It used to accept an unauthenticated
    POST from anyone and mark any order fully paid."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    order = get_object_or_404(Order, id=order_id)
    if not _can_view_order(request, order):
        return JsonResponse({'success': False, 'error': 'Not allowed'}, status=403)
    if request.user.is_authenticated and request.user.is_staff:
        order.payment_status = 'paid_online'
    else:
        order.payment_status = 'online_pending'
    order.payment_method = 'online'
    order.save()
    return JsonResponse({'success': True, 'payment_status': order.payment_status})


def download_bill(request, order_id):
    """Keep for backward compat but redirect to bill view"""
    return redirect('bill', order_id=order_id)




def print_bill(request, order_id):
    """Thermal printer-friendly bill page for staff"""
    order = get_object_or_404(Order, id=order_id)
    if not _can_view_order(request, order):
        return _deny_order(request)
    shop = get_shop()
    root_order = order
    while root_order.parent_order:
        root_order = root_order.parent_order

    reorder_only = request.GET.get('reorder_only') == '1'

    if reorder_only and order.parent_order:
        # Print only the REORDER items (not the full bill)
        combined_items = []
        combined_subtotal = Decimal('0')
        for item in order.items.select_related('menu_item', 'combo').prefetch_related('combo__combo_items__menu_item').all():
            combined_items.append({
                'name': item.display_name,
                'is_combo': item.is_combo,
                'is_free': item.is_free,
                'combo_items': item.combo_components,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'total_price': item.total_price,
                'notes': item.notes,
            })
            combined_subtotal += item.total_price
        discount_percent = order.discount_percent
        combined_parcel  = order.parcel_charge or Decimal('0')
        if order.discount_is_flat:
            discount_amount = min(order.discount_amount or Decimal('0'), combined_subtotal)
        else:
            discount_amount = (combined_subtotal * discount_percent / 100).quantize(Decimal('0.01')) if discount_percent else Decimal('0')
        combined_total = combined_subtotal + combined_parcel - discount_amount
        bill_order = order  # use the reorder itself
    else:
        # Full bill: one section per order in the chain, each with its own items
        # and its own offer/discount. This matches the website bill layout.
        all_chain = [root_order] + list(root_order.reorders.order_by('created_at'))
        active_chain = [o for o in all_chain if o.status != 'cancelled']
        order_sections = []
        combined_subtotal = Decimal('0')
        discount_amount   = Decimal('0')
        combined_parcel   = Decimal('0')
        combined_total    = Decimal('0')
        for o in active_chain:
            items_qs = list(o.items.select_related('menu_item','combo').prefetch_related('combo__combo_items__menu_item').all())
            sub = sum(i.total_price for i in items_qs)
            par = o.parcel_charge or Decimal('0')
            if o.discount_is_flat:
                disc = min(o.discount_amount or Decimal('0'), sub)
            elif o.discount_percent:
                disc = (sub * o.discount_percent / 100).quantize(Decimal('0.01'))
            else:
                disc = Decimal('0')
            sec_total = max(Decimal('0'), sub + par - disc)
            combined_subtotal += sub
            combined_parcel   += par
            discount_amount   += disc
            combined_total    += sec_total
            order_sections.append({
                'order_number': o.order_number,
                'is_reorder': o.parent_order_id is not None,
                'items': [{
                    'name': item.display_name,
                    'is_combo': item.is_combo,
                    'is_free': item.is_free,
                    'combo_items': item.combo_components,
                    'quantity': item.quantity,
                    'unit_price': item.unit_price,
                    'total_price': item.total_price,
                    'notes': item.notes,
                } for item in items_qs],
                'subtotal': sub,
                'parcel': par,
                'discount': disc,
                'offer_title': o.offer_title or '',
                'discount_percent': o.discount_percent,
                'discount_is_flat': o.discount_is_flat,
                'section_total': sec_total,
            })
        # flat list for templates that don't use sections
        combined_items = [item for s in order_sections for item in s['items']]
        discount_percent = root_order.discount_percent
        bill_order = root_order
    has_reorders = len(active_chain) > 1 if not (reorder_only and order.parent_order) else False
    return render(request, 'restaurant/print_bill.html', {
        'order': bill_order,
        'root_order': root_order,
        'is_reorder_only': reorder_only and order.parent_order is not None,
        'order_sections': order_sections if not (reorder_only and order.parent_order) else None,
        'has_reorders': has_reorders,
        'combined_items': combined_items,
        'combined_subtotal': combined_subtotal,
        'combined_parcel': combined_parcel,
        'discount_percent': discount_percent,
        'discount_amount': discount_amount,
        'discount_is_flat': root_order.discount_is_flat,
        'offer_title': root_order.offer_title or '',
        'combined_total': combined_total,
        'shop': shop,
    })



def reorder_notification_api(request):
    """Returns merged-reorder notifications for staff sound alert.
    Reorders now merge items into the parent order directly, so we use
    a cache queue instead of querying child orders."""
    try:
        import time as _time
        last_ts = int(request.GET.get('last_ts', 0))
        from django.core.cache import cache
        notifs = cache.get('staff_reorder_notifs', [])
        # Return only notifications newer than last_ts
        result = [n for n in notifs if n.get('ts', 0) > last_ts]
        return JsonResponse({'orders': result, 'server_ts': int(_time.time())})
    except Exception as e:
        return JsonResponse({'orders': [], 'error': str(e)})

# --- STAFF PORTAL -------------------------------------------------

def staff_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user and user.is_staff:
            login(request, user)
            return redirect('staff_portal')
        return render(request, 'restaurant/staff_login.html', {'error': 'Invalid credentials'})
    return render(request, 'restaurant/staff_login.html')


def staff_logout(request):
    logout(request)
    return redirect('staff_login')



# ===================================================================
# LIVE ORDER PAGE
# ===================================================================

def live_order(request):
    """Dedicated live order page - reads order_id from query param or session."""
    if not request.session.get('customer_phone'):
        return redirect('/customer/login/')
    return render(request, 'restaurant/live_order.html', {})


def live_order_data(request):
    """JSON API for the live order page - returns full order state."""
    order_id = request.GET.get('order_id')
    if not order_id:
        return JsonResponse({'error': 'no order id'}, status=400)
    try:
        root = Order.objects.get(id=order_id)
        # Walk to root
        while root.parent_order:
            root = root.parent_order
    except Order.DoesNotExist:
        return JsonResponse({'error': 'not found'}, status=404)
    if not _can_view_order(request, root):
        return JsonResponse({'error': 'Not allowed'}, status=403)

    all_orders = [root] + list(root.reorders.order_by('created_at'))
    all_completed = all(o.status == 'completed' for o in all_orders)

    # Include all non-cancelled child orders (including 'ready' ones).
    # Items are no longer merged into root, so each order holds its own items.
    active_children = [o for o in all_orders[1:] if o.status not in ('completed','cancelled')]
    display_orders = [root] + active_children

    orders_data = []
    for o in display_orders:
        items_data = []
        for item in o.items.select_related('menu_item', 'combo').prefetch_related('combo__combo_items__menu_item').all():
            items_data.append({
                'name': item.display_name,
                'notes': item.notes or '',
                'is_combo': item.is_combo,
                'is_free': item.is_free,
                'combo_items': item.combo_components,
                'qty':  item.quantity,
                'total': float(item.total_price),
            })
        orders_data.append({'items': items_data, 'total': float(o.total_amount)})

    # Sum totals only from display_orders (root already includes merged items)
    combined_subtotal = sum(o.subtotal for o in display_orders)
    combined_parcel   = sum((o.parcel_charge or Decimal('0')) for o in display_orders)
    combined_discount = sum((o.discount_amount or Decimal('0')) for o in display_orders)
    # Compute per-order totals and sum them for an accurate grand total.
    # Using root.total_amount alone misses pending/ready reorders that
    # haven't been stored there yet.
    def _order_total_live(o):
        sub = o.subtotal or Decimal('0')
        if o.discount_is_flat:
            disc = min(o.discount_amount or Decimal('0'), sub)
        elif o.discount_percent:
            disc = (sub * o.discount_percent / 100).quantize(Decimal('0.01'))
        else:
            disc = Decimal('0')
        return max(Decimal('0'), sub + (o.parcel_charge or Decimal('0')) - disc)
    combined_total = sum(_order_total_live(o) for o in display_orders)
    # Fall back to root.total_amount when it's already the correct session total
    # (e.g. after the root is complete).
    if root.status == 'completed':
        combined_total = root.total_amount

    return JsonResponse({
        'order_id':        root.id,
        'order_number':    root.order_number,
        'status':          root.status,
        'all_completed':   all_completed,
        'customer_name':   root.customer_name,
        'customer_phone':  root.customer_phone,
        'combined_total':  float(combined_total),
        'combined_parcel': float(combined_parcel),
        'orders':          orders_data,
    })


def combo_detail_api(request, combo_id):
    """JSON API - returns combo details with items for customer modal."""
    try:
        combo = Combo.objects.prefetch_related('combo_items__menu_item').get(id=combo_id, is_active=True)
    except Combo.DoesNotExist:
        return JsonResponse({'error': 'not found'}, status=404)
    items_data = []
    for ci in combo.combo_items.all():
        mi = ci.menu_item
        try:
            mi_image = mi.image.url if mi.image else None
        except Exception:
            mi_image = None
        items_data.append({
            'name': mi.name,
            'qty': ci.quantity,
            'price': float(mi.price),
            'image': mi_image,
            'type': mi.item_type,
        })
    try:
        combo_image = combo.image.url if combo.image else None
    except Exception:
        combo_image = None
    return JsonResponse({
        'id': combo.id,
        'name': combo.name,
        'description': combo.description,
        'price': float(combo.price),
        'icon': combo.icon or '🎁',
        'image': combo_image,
        'items': items_data,
    })



# -- POS DRAFT VIEWS --------------------------------------------

@login_required(login_url='/staff/login/')
def pos_save_draft(request):
    """Save a POS draft order to the database (includes slot, parcel, order_type)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        import json as _json
        from decimal import Decimal as D
        data = _json.loads(request.body)
        name       = (data.get('name') or 'Walk-in').strip()
        phone      = (data.get('phone') or '').strip()
        slot       = (data.get('slot') or '').strip()
        table      = (data.get('table') or slot or 'Takeaway').strip()
        items      = data.get('items', [])
        sub        = D(str(data.get('subtotal', 0)))
        parcel     = D(str(data.get('parcel', 0)))
        disc       = D(str(data.get('discount_pct', 0)))
        total      = D(str(data.get('total', 0)))
        note       = (data.get('note') or '').strip()
        order_type = (data.get('order_type') or 'dine_in').strip()
        draft_id   = data.get('draft_id')
        table_name = slot if slot else table

        if draft_id:
            try:
                draft = PosDraft.objects.get(id=draft_id)
                draft.customer_name  = name
                draft.customer_phone = phone
                draft.table_name     = table_name
                draft.items_json     = _json.dumps(items)
                draft.subtotal       = sub
                draft.discount_pct   = disc
                draft.total_amount   = total
                draft.note           = note
                # Store extra fields in note if model doesn't have them
                draft.save()
            except PosDraft.DoesNotExist:
                draft_id = None

        if not draft_id:
            import random
            num = f"D{timezone.now().strftime('%Y%m%d')}{random.randint(100,999)}"
            draft = PosDraft.objects.create(
                draft_number=num, customer_name=name, customer_phone=phone,
                table_name=table_name, items_json=_json.dumps(items),
                subtotal=sub, discount_pct=disc, total_amount=total, note=note
            )

        return JsonResponse({
            'success': True,
            'draft_id': draft.id,
            'draft_number': draft.draft_number,
            'slot': slot,
            'order_type': order_type,
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def pos_get_drafts(request):
    """Return all active POS drafts (includes slot and order_type derived from table_name)."""
    import json as _json
    drafts = PosDraft.objects.filter(is_deleted=False).order_by('-created_at')[:50]
    result = []
    for d in drafts:
        table_name = d.table_name or 'Takeaway'
        is_tw = table_name == 'Takeaway' or table_name.startswith('Takeaway ')
        try:
            raw_items = _json.loads(d.items_json)
        except Exception:
            raw_items = []
        result.append({
            'id': d.id,
            'draft_number': d.draft_number,
            'customer_name': d.customer_name,
            'customer_phone': d.customer_phone,
            'table_name': table_name,
            'slot': table_name,
            'order_type': 'takeaway' if is_tw else 'dine_in',
            'total_amount': float(d.total_amount),
            'subtotal': float(d.subtotal),
            'parcel': 0,  # parcel not stored separately in current schema
            'items': raw_items,
            'discount_pct': float(d.discount_pct),
            'note': d.note,
            'created_at': d.created_at.strftime('%d %b %Y %H:%M'),
        })
    return JsonResponse({'drafts': result})


@login_required(login_url='/staff/login/')
def pos_delete_draft(request, draft_id):
    """Soft-delete a POS draft."""
    try:
        d = PosDraft.objects.get(id=draft_id)
        d.is_deleted = True
        d.save()
        return JsonResponse({'success': True})
    except PosDraft.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Not found'}, status=404)


def pos_save_draft_no_auth(request):
    """Save a POS draft order to the database without authentication requirement."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        import json as _json
        from decimal import Decimal as D
        data = _json.loads(request.body)
        name       = (data.get('name') or 'Walk-in').strip()
        phone      = (data.get('phone') or '').strip()
        slot       = (data.get('slot') or '').strip()
        table      = (data.get('table') or slot or 'Takeaway').strip()
        items      = data.get('items', [])
        sub        = D(str(data.get('subtotal', 0)))
        parcel     = D(str(data.get('parcel', 0)))
        disc       = D(str(data.get('discount_pct', 0)))
        total      = D(str(data.get('total', 0)))
        note       = (data.get('note') or '').strip()
        order_type = (data.get('order_type') or 'dine_in').strip()
        draft_id   = data.get('draft_id')
        table_name = slot if slot else table

        if draft_id:
            try:
                draft = PosDraft.objects.get(id=draft_id)
                draft.customer_name  = name
                draft.customer_phone = phone
                draft.table_name     = table_name
                draft.items_json     = _json.dumps(items)
                draft.subtotal       = sub
                draft.discount_pct   = disc
                draft.total_amount   = total
                draft.note           = note
                draft.save()
            except PosDraft.DoesNotExist:
                draft_id = None

        if not draft_id:
            import random
            num = f"D{timezone.now().strftime('%Y%m%d')}{random.randint(100,999)}"
            draft = PosDraft.objects.create(
                draft_number=num, customer_name=name, customer_phone=phone,
                table_name=table_name, items_json=_json.dumps(items),
                subtotal=sub, discount_pct=disc, total_amount=total, note=note
            )

        return JsonResponse({'success': True, 'draft_id': draft.id, 'draft_number': draft.draft_number})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def pos_cart_live_push(request):
    """Receives the full POS cart (all slots) and upserts PosDraft records
    so the staff portal live sidebar can display them in real time.
    Called from POS terminal every time items change (addItem / removeItem).
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        import json as _json
        data = _json.loads(request.body)
        slots = data.get('slots', {})   # { "Table 1": {items, disc, isTW, parcel, name}, ... }

        # Mark sentinel so we can tell live-cart drafts from real drafts
        LIVE_MARKER = '__LIVE__'

        for slot_name, cart in slots.items():
            items = cart.get('items', [])
            # Build subtotal
            sub = sum(i.get('price', 0) * i.get('qty', 1) for i in items)
            disc_pct = float(cart.get('disc', 0) or 0)
            parcel = float(cart.get('parcel', 0) or 0)
            dis = round(sub * disc_pct / 100, 2)
            total = round(sub + parcel - dis, 2)

            # Upsert: find existing live-cart draft for this slot, or create
            draft = PosDraft.objects.filter(
                table_name=slot_name,
                note__startswith=LIVE_MARKER,
                is_deleted=False
            ).first()

            items_json = _json.dumps(items)
            if draft:
                draft.items_json   = items_json
                draft.subtotal     = sub
                draft.discount_pct = disc_pct
                draft.total_amount = total
                draft.customer_name = cart.get('name') or 'Walk-in'
                draft.save()
            else:
                import random
                num = f"LC{timezone.now().strftime('%Y%m%d%H%M%S')}{random.randint(10,99)}"
                PosDraft.objects.create(
                    draft_number=num,
                    customer_name=cart.get('name') or 'Walk-in',
                    customer_phone=cart.get('phone') or '',
                    table_name=slot_name,
                    items_json=items_json,
                    subtotal=sub,
                    discount_pct=disc_pct,
                    total_amount=total,
                    note=f"{LIVE_MARKER}{slot_name}",
                )

        # Delete live-cart drafts for slots that are now empty / removed
        active_slots = set(slots.keys())
        PosDraft.objects.filter(
            note__startswith=LIVE_MARKER,
            is_deleted=False
        ).exclude(table_name__in=active_slots).update(is_deleted=True)

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def pos_save_order(request):
    """Save a POS order directly into the staff portal Order table.
    No customer session required. Staff login required.
    Status = completed + paid so it goes straight to Excel, no staff portal popup.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        import random as _random
        data = json.loads(request.body)

        name         = (data.get('customer_name') or 'Walk-in').strip() or 'Walk-in'
        phone        = (data.get('customer_phone') or '0000000000').strip() or '0000000000'
        items_data   = data.get('items', [])
        disc_pct     = Decimal(str(data.get('discount_percent', 0) or 0))
        order_type   = (data.get('order_type') or 'dine_in').strip()
        slot         = (data.get('slot') or '').strip()
        parcel_total = Decimal(str(data.get('parcel_total', 0) or 0))
        offer_flat   = Decimal(str(data.get('cart_offer_flat', 0) or 0))
        offer_title  = (data.get('cart_offer_title') or '').strip()

        if not items_data:
            return JsonResponse({'success': False, 'error': 'No items in order'})

        # Pre-fetch all regular menu items in one query
        regular_ids = [int(d['id']) for d in items_data if d.get('id')]
        menu_items_map = {m.id: m for m in MenuItem.objects.filter(id__in=regular_ids)} if regular_ids else {}

        # Pre-fetch combo items
        combo_ids = [int(d['combo_id']) for d in items_data if d.get('combo_id')]
        combos_map = {c.id: c for c in Combo.objects.filter(id__in=combo_ids, is_active=True)} if combo_ids else {}

        # Create the order
        order = Order(
            customer_name=name,
            customer_phone=phone,
            special_instructions=f'[POS][Slot:{slot}]' if slot else '[POS]',
            discount_percent=disc_pct,
            order_type=order_type,
            order_number=f"POS{timezone.now().strftime('%Y%m%d')}{_random.randint(1000, 9999)}",
            subtotal=Decimal('0'),
            discount_amount=Decimal('0'),
            total_amount=Decimal('0'),
            status='completed',
            payment_status='paid',
        )
        order.save()

        subtotal = Decimal('0')
        to_create = []

        for d in items_data:
            qty        = max(1, int(d.get('qty', 1)))
            sent_price = Decimal(str(d.get('price', 0) or 0))

            if d.get('combo_id'):
                # Combo item - use combo's representative first item or sent price
                combo = combos_map.get(int(d['combo_id']))
                if not combo:
                    # No DB record - create a placeholder using first available item
                    fallback = MenuItem.objects.filter(is_available=True).first()
                    if not fallback:
                        continue
                    unit_price = sent_price if sent_price > 0 else Decimal('0')
                    to_create.append(OrderItem(
                        order=order,
                        menu_item=fallback,
                        quantity=qty,
                        unit_price=unit_price,
                        notes=f'Combo: {d.get("name", "?")}',
                    ))
                    subtotal += unit_price * qty
                    continue
                # Use first combo item's menu_item as the representative row
                combo_lines = list(combo.combo_items.select_related('menu_item').all())
                rep_item = combo_lines[0].menu_item if combo_lines else MenuItem.objects.filter(is_available=True).first()
                if not rep_item:
                    continue
                # Always price a combo from the DB, never from the terminal.
                unit_price = combo.price
                to_create.append(OrderItem(
                    order=order,
                    menu_item=rep_item,
                    combo=combo,
                    quantity=qty,
                    unit_price=unit_price,
                    notes=', '.join(f'{ci.quantity}x {ci.menu_item.name}' for ci in combo_lines),
                ))
                subtotal += unit_price * qty

            else:
                # Regular item
                mid = d.get('id')
                if not mid:
                    continue
                menu_item = menu_items_map.get(int(mid))
                if not menu_item:
                    continue
                is_free_line = bool(d.get('free_offer'))
                if is_free_line:
                    # A free-item cart offer. Previously `price 0` was read as
                    # "nothing sent" and the item was billed at full price.
                    unit_price = Decimal('0')
                elif d.get('from_offer') and Decimal('0') <= sent_price <= menu_item.price:
                    unit_price = sent_price
                else:
                    unit_price = sent_price if sent_price > 0 else menu_item.price
                to_create.append(OrderItem(
                    order=order,
                    menu_item=menu_item,
                    quantity=qty,
                    unit_price=unit_price,
                    is_free=is_free_line,
                    notes='🎁 FREE with offer' if is_free_line else '',
                ))
                subtotal += unit_price * qty

        OrderItem.objects.bulk_create(to_create)

        is_flat = False
        if offer_flat > 0 and subtotal > 0:
            disc_amount = min(offer_flat, subtotal).quantize(Decimal('0.01'))
            disc_pct = Decimal('0')
            is_flat = True
        elif disc_pct:
            disc_amount = (subtotal * disc_pct / 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            disc_amount = Decimal('0')
        total = subtotal + parcel_total - disc_amount
        if total < 0:
            total = Decimal('0')

        Order.objects.filter(pk=order.pk).update(
            subtotal=subtotal,
            discount_percent=disc_pct,
            discount_amount=disc_amount,
            discount_is_flat=is_flat,
            offer_title=offer_title,
            total_amount=total,
            parcel_charge=parcel_total,
        )

        # Clean up the live-cart draft for this slot so sidebar removes it
        if slot:
            PosDraft.objects.filter(
                table_name=slot,
                note__startswith='__LIVE__',
                is_deleted=False
            ).update(is_deleted=True)

        return JsonResponse({
            'success': True,
            'order_id': order.id,
            'order_number': order.order_number,
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def staff_portal(request):
    from django.db.models import Count, Avg
    tables = Table.objects.filter(is_active=True).order_by('number')
    pending_orders  = Order.objects.filter(status='pending').select_related('table').prefetch_related('items__menu_item').order_by('created_at')
    accepted_orders = Order.objects.filter(status='accepted').select_related('table').prefetch_related('items__menu_item').order_by('created_at')
    cooking_orders  = Order.objects.filter(status='preparing').select_related('table').prefetch_related('items__menu_item').order_by('created_at')
    ready_orders    = Order.objects.filter(status='ready', parent_order__isnull=True).select_related('table').prefetch_related('items__menu_item').order_by('created_at')
    completed_orders = Order.objects.filter(status='completed').select_related('table').prefetch_related('items__menu_item').order_by('-created_at')[:30]

    def _fix_totals(qs):
        # For each order: recalculate from own items + own discount.
        # For ROOT orders: also add the totals of any ready/completed reorders
        # so the session total shown in the combined bill is correct.
        result = list(qs)
        for o in result:
            items = list(o.items.all())
            sub = sum((i.unit_price * i.quantity for i in items), Decimal('0'))
            if o.discount_is_flat:
                disc = min(o.discount_amount or Decimal('0'), sub)
            elif o.discount_percent:
                disc = (sub * o.discount_percent / 100).quantize(Decimal('0.01'))
            else:
                disc = Decimal('0')
            parcel = o.parcel_charge or Decimal('0')
            own_total = max(Decimal('0'), sub + parcel - disc)
            # Root orders: add ready + completed reorder totals so that
            # "Total Bill" on the combined bill card is the session total.
            if not o.parent_order_id:
                reorder_total = sum(
                    r.total_amount
                    for r in o.reorders.filter(status__in=['ready', 'completed'])
                )
                o.total_amount = own_total + reorder_total
            else:
                o.total_amount = own_total
        return result

    pending_orders  = _fix_totals(pending_orders)
    accepted_orders = _fix_totals(accepted_orders)
    cooking_orders  = _fix_totals(cooking_orders)
    ready_orders    = _fix_totals(ready_orders)
    completed_orders = _fix_totals(completed_orders)

    # Build combined bill info for ready orders, then pair with each order
    ready_order_meta = {}
    for o in ready_orders:
        root = o
        while root.parent_order_id:
            root = root.parent_order
        all_linked = get_full_order_chain(root)
        active_linked = get_active_order_chain(root)
        has_reorders = len(all_linked) > 1
        if has_reorders:
            # Reorders are now shown separately in the Ready column, but
            # the root card's combined bill must include their discounts.
            reorders = [x for x in all_linked if x.parent_order_id and x.status != 'cancelled']
            ready_order_meta[o.id] = {
                'has_reorders': True,
                'combined_total': o.total_amount,  # already set to session total by _fix_totals
                'root_order_number': root.order_number,
                'reorder_count': len(reorders),
                'reorders_detail': reorders,
            }
        else:
            ready_order_meta[o.id] = {
                'has_reorders': False,
                'combined_total': None,
                'root_order_number': None,
                'reorder_count': 0,
            }

    # Pair each ready order with its combined bill meta for easy template access
    ready_orders_with_meta = [
        (o, ready_order_meta.get(o.id, {'has_reorders': False, 'combined_total': None, 'reorder_count': 0, 'root_order_number': None}))
        for o in ready_orders
    ]

    today = timezone.localdate()   # use local date (IST), not UTC
    week_start  = today - timezone.timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    def _revenue_from_items(order_qs):
        # root.total_amount now holds the full session total (root + reorders)
        # after fix 2 stores it on completion. Simply sum the stored values.
        return sum(o.total_amount for o in order_qs)

    # Today stats - only ROOT orders (parent_order__isnull=True) to avoid double-counting
    # reorder child orders: their items are merged into the root when marked ready,
    # so counting child orders separately inflates revenue.
    # Use a timezone-aware range so midnight-IST orders are not missed
    # (SQLite's __date lookup works in UTC, which is 5h30m behind IST).
    import datetime as _dt
    _tz = timezone.get_current_timezone()
    _today_start = timezone.make_aware(_dt.datetime.combine(today, _dt.time.min), _tz)
    _today_end   = timezone.make_aware(_dt.datetime.combine(today, _dt.time.max), _tz)

    today_qs = Order.objects.filter(status='completed', parent_order__isnull=True,
                                    created_at__range=(_today_start, _today_end)).prefetch_related('items')
    today_revenue       = _revenue_from_items(today_qs)
    today_orders_count  = Order.objects.filter(
        created_at__range=(_today_start, _today_end),
        parent_order__isnull=True).count()
    today_completed     = today_qs.count()

    # Weekly stats
    _week_start_dt = timezone.make_aware(_dt.datetime.combine(week_start, _dt.time.min), _tz)
    week_qs      = Order.objects.filter(status='completed', parent_order__isnull=True,
                                        created_at__gte=_week_start_dt).prefetch_related('items')
    week_revenue = _revenue_from_items(week_qs)
    week_orders  = week_qs.count()

    _month_start_dt = timezone.make_aware(_dt.datetime.combine(month_start, _dt.time.min), _tz)

    # Monthly stats
    month_qs      = Order.objects.filter(status='completed', parent_order__isnull=True, created_at__gte=_month_start_dt).prefetch_related('items')
    month_revenue = _revenue_from_items(month_qs)
    month_orders  = month_qs.count()

    # All-time stats
    all_qs        = Order.objects.filter(status='completed', parent_order__isnull=True).prefetch_related('items')
    all_revenue   = _revenue_from_items(all_qs)
    all_orders    = all_qs.count()
    avg_order_val = (all_revenue / all_orders) if all_orders else Decimal('0')

    # Top 8 selling items (by quantity sold this month)
    # F is imported at top as F_db
    top_items = (
        OrderItem.objects
        .filter(order__status='completed', order__parent_order__isnull=True, order__created_at__date__gte=month_start)
        .values(item_name=F_db('menu_item__name'))
        .annotate(
            total_qty=Sum('quantity'),
            total_rev=Sum(
                ExpressionWrapper(
                    F_db('unit_price') * F_db('quantity'),
                    output_field=DecimalField()
                )
            )
        )
        .order_by('-total_qty')[:8]
    )

    # Last 7 days revenue for chart
    from datetime import timedelta
    chart_labels = []
    chart_values = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        day_orders = Order.objects.filter(status='completed', parent_order__isnull=True, created_at__date=d).prefetch_related('items')
        rev = _revenue_from_items(day_orders)
        chart_labels.append(d.strftime('%d %b'))
        chart_values.append(float(rev))

    shop = get_shop()
    discounts = Discount.objects.filter(is_active=True)
    pending_ids = [o.id for o in pending_orders]
    all_menu_items = list(
        MenuItem.objects.filter(is_available=True).select_related('category').order_by('category__name', 'name')
        .values('id', 'name', 'price', 'category__name')
    )

    return render(request, 'restaurant/staff_portal.html', {
        'tables': tables,
        'pending_orders': pending_orders,
        'accepted_orders': accepted_orders,
        'cooking_orders': cooking_orders,
        'ready_orders': ready_orders,
        'ready_order_meta': ready_order_meta,
        'ready_orders_with_meta': ready_orders_with_meta,
        'completed_orders': completed_orders,
        'today_revenue': today_revenue,
        'today_orders_count': today_orders_count,
        'today_completed': today_completed,
        'week_revenue': week_revenue,
        'week_orders': week_orders,
        'month_revenue': month_revenue,
        'month_orders': month_orders,
        'all_revenue': all_revenue,
        'all_orders': all_orders,
        'avg_order_val': avg_order_val,
        'top_items': top_items,
        'chart_labels_json': json.dumps(chart_labels),
        'chart_values_json': json.dumps(chart_values),
        'shop': shop,
        'discounts': discounts,
        'pending_ids_json': json.dumps(pending_ids),
        'all_menu_items_json': json.dumps([
            {'id': m['id'], 'name': m['name'], 'price': float(m['price']), 'category': m['category__name'] or 'Other'}
            for m in all_menu_items
        ]),
    })


@login_required(login_url='/staff/login/')
@require_POST
def update_order_status(request, order_id):
    try:
        with transaction.atomic():
            order = Order.objects.select_for_update().get(pk=order_id)
            data = json.loads(request.body)
            new_status = data.get('status')
            payment_method = data.get('payment_method', '')
            discount_percent = data.get('discount_percent')
            cash_received = data.get('cash_received')
            change_amount = data.get('change_amount')

            # Discount handling on completion.
            #   None / ''  -> leave the order exactly as the customer left it
            #   > 0        -> staff override, replaces any customer offer
            #   -1         -> staff explicitly removes every discount
            # Previously a plain 0 was sent by default, which silently wiped the
            # customer's percentage cart offer and pushed the total back up to
            # the undiscounted amount at the moment of billing.
            if discount_percent is not None and str(discount_percent).strip() != '':
                new_pct = Decimal(str(discount_percent))
                if new_pct > 0:
                    order.discount_percent = new_pct
                    order.discount_is_flat = False
                    order.offer_title = ''
                elif new_pct < 0:
                    order.discount_percent = Decimal('0')
                    order.discount_amount = Decimal('0')
                    order.discount_is_flat = False
                    order.offer_title = ''
                # new_pct == 0 -> no change; the customer's offer stands.

            if cash_received is not None:
                order.cash_received = Decimal(str(cash_received))
            if change_amount is not None:
                order.change_amount = Decimal(str(change_amount))

            # Reorders (child orders with a parent) go directly to preparing when accepted
            if order.status == 'pending' and order.parent_order_id and new_status == 'preparing':
                pass  # allowed

            # If this child reorder was already merged, return immediately without re-merging.
            if new_status == 'ready' and order.parent_order_id and order.status == 'completed':
                root = order.parent_order
                while root.parent_order_id:
                    root = root.parent_order
                return JsonResponse({
                    'success': True,
                    'status': order.status,
                    'reorder_merged': True,
                    'root_id': root.id,
                    'root_total': float(root.total_amount),
                })

            valid_transitions = {
                'pending':   ['accepted', 'preparing', 'cancelled'],
                'accepted':  ['preparing', 'cancelled'],
                'preparing': ['ready', 'cancelled'],
                'ready':     ['completed', 'cancelled'],
                'completed': [], 'cancelled': [],
            }
            if new_status not in valid_transitions.get(order.status, []):
                return JsonResponse({'success': False, 'error': f'Cannot move from {order.status} to {new_status}'})

            # Each reorder keeps its OWN items and discount in its own Order row.
            # Items are NOT merged into the root — each order stands alone.
            if new_status == 'ready' and order.parent_order_id:
                # Recalculate this reorder's own totals.
                order.calculate_totals()
                # Keep status as 'ready' so staff can still see it in the Ready
                # column — it disappears only when the root is completed.
                Order.objects.filter(pk=order.pk).update(
                    status='ready',
                    subtotal=order.subtotal,
                    discount_amount=order.discount_amount,
                    total_amount=order.total_amount,
                )
                root = order.parent_order
                while root.parent_order_id:
                    root = root.parent_order
                # Store the session total on root so the combined bill is correct.
                chain = [o for o in _get_chain(root) if o.status != 'cancelled']
                session_total = sum(o.total_amount for o in chain)
                Order.objects.filter(pk=root.pk).update(total_amount=session_total)
                return JsonResponse({
                    'success': True,
                    'status': 'ready',
                    'reorder_merged': True,
                    'root_id': root.id,
                    'root_total': float(session_total),
                    'has_reorders': True,
                    'combined_total': float(session_total),
                })

            order.status = new_status
            if new_status == 'completed':
                order.completed_at = timezone.now()
                if payment_method == 'offline':
                    order.payment_status = 'paid_offline'
                    order.payment_method = 'offline'
                elif payment_method == 'online':
                    order.payment_status = 'paid_online'
                    order.payment_method = 'online'
                if not order.parent_order_id:
                    order.cash_received = None
                    order.change_amount = None
            order.calculate_totals()
            # Persist totals directly so they are never overwritten by model.save() re-calc
            Order.objects.filter(pk=order.pk).update(
                status=order.status,
                completed_at=order.completed_at,
                payment_status=order.payment_status,
                payment_method=order.payment_method,
                discount_percent=order.discount_percent,
                discount_is_flat=order.discount_is_flat,
                offer_title=order.offer_title,
                subtotal=order.subtotal,
                parcel_charge=order.parcel_charge,
                discount_amount=order.discount_amount,
                total_amount=order.total_amount,
            )
            # After completing a root order, auto-complete its ready reorders
            # and overwrite total_amount with the full SESSION total so that
            # revenue, stats and the customer bill all show the correct figure.
            if new_status == 'completed' and not order.parent_order_id:
                order.reorders.filter(status='ready').update(
                    status='completed',
                    completed_at=timezone.now(),
                    payment_status=order.payment_status or 'paid_offline',
                )
                chain_orders = [order] + list(order.reorders.filter(status='completed'))
                session_rev = sum(o.total_amount for o in chain_orders)
                # Overwrite AFTER the main update above so nothing resets it.
                Order.objects.filter(pk=order.pk).update(total_amount=session_rev)
                order.total_amount = session_rev
            # Also call save() so table status logic runs (skip_recalc=True avoids overwriting our update above)
            order.save(skip_recalc=True)

            # When marking ready or completing, include reorder info
        extra = {}
        if new_status in ('ready', 'completed'):
            root_order = order
            while root_order.parent_order:
                root_order = root_order.parent_order
            all_related = get_active_order_chain(root_order)
            has_reorders = len(all_related) > 1
            combined_total = sum(float(o.total_amount) for o in all_related)
            extra['has_reorders'] = has_reorders
            extra['root_id'] = root_order.id
            extra['combined_total'] = round(combined_total, 2) if has_reorders else None
        return JsonResponse({'success': True, 'status': order.status, **extra})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def staff_order_items_get(request, order_id):
    """Return current items of an order plus full menu for the edit modal."""
    try:
        order = get_object_or_404(Order, id=order_id)
        items = []
        for oi in order.items.select_related('menu_item', 'combo').all():
            items.append({
                'order_item_id': oi.id,
                'menu_item_id': oi.menu_item.id,
                'name': oi.display_name,
                'is_combo': oi.is_combo,
                'is_free': oi.is_free,
                'quantity': oi.quantity,
                'unit_price': float(oi.unit_price),
                'total': float(oi.unit_price * oi.quantity),
                'notes': oi.notes or '',
            })
        menu_qs = MenuItem.objects.filter(is_available=True).select_related('category').order_by('category__name', 'name')
        menu = []
        for m in menu_qs:
            menu.append({
                'id': m.id,
                'name': m.name,
                'price': float(m.price),
                'category': m.category.name if m.category else 'Other',
            })
        return JsonResponse({
            'success': True,
            'order_id': order.id,
            'order_number': order.order_number,
            'items': items,
            'menu': menu,
            'total': float(order.total_amount),
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def staff_edit_order_items(request, order_id):
    """Add or remove items on a cooking/ready order from staff portal."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        order = get_object_or_404(Order, id=order_id)
        data = json.loads(request.body)
        # remove_ids: list of OrderItem IDs to delete
        remove_ids = data.get('remove_ids', [])
        # add_items: list of {menu_item_id, quantity}
        add_items = data.get('add_items', [])
        # update_items: list of {order_item_id, quantity}
        update_items = data.get('update_items', [])

        if remove_ids:
            OrderItem.objects.filter(id__in=remove_ids, order=order).delete()

        for upd in update_items:
            oi = OrderItem.objects.filter(id=upd['order_item_id'], order=order).first()
            if oi:
                qty = int(upd.get('quantity', 1))
                if qty <= 0:
                    oi.delete()
                else:
                    oi.quantity = qty
                    oi.save()

        for ai in add_items:
            menu_item = MenuItem.objects.filter(id=ai["menu_item_id"], is_available=True).first()
            if not menu_item:
                continue
            qty = max(1, int(ai.get('quantity', 1)))
            existing = order.items.filter(menu_item=menu_item).first()
            if existing:
                existing.quantity += qty
                existing.save()
            else:
                OrderItem.objects.create(order=order, menu_item=menu_item,
                                         quantity=qty, unit_price=menu_item.price)

        order.calculate_totals()
        Order.objects.filter(pk=order.pk).update(
            subtotal=order.subtotal,
            parcel_charge=order.parcel_charge,
            discount_amount=order.discount_amount,
            total_amount=order.total_amount,
        )
        return JsonResponse({'success': True, 'new_total': float(order.total_amount)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def pending_orders_api(request):
    """Polling endpoint - returns new pending orders since last_id"""
    try:
        last_id = int(request.GET.get('last_id', 0))
        # Include reorders (child orders) - they show as pending for staff to accept
        orders = Order.objects.filter(id__gt=last_id, status='pending').select_related('table', 'parent_order')
        result = []
        for o in orders:
            # Compute total on-the-fly if stored value is 0 (can happen
            # immediately after order creation before calculate_totals saves).
            total = float(o.total_amount)
            if total == 0:
                from decimal import Decimal as _D
                items = list(o.items.all())
                sub = sum(i.unit_price * i.quantity for i in items)
                if o.discount_is_flat:
                    disc = min(o.discount_amount or _D('0'), sub)
                elif o.discount_percent:
                    disc = (sub * o.discount_percent / 100).quantize(_D('0.01'))
                else:
                    disc = _D('0')
                total = float(max(_D('0'), sub + (o.parcel_charge or _D('0')) - disc))
            result.append({
                'id': o.id,
                'order_number': o.order_number,
                'customer_name': o.customer_name,
                'table': o.table.name if o.table else 'Takeaway',
                'total': total,
                'order_type': o.order_type,
                'special_instructions': o.special_instructions or '',
                'is_reorder': o.parent_order_id is not None,
                'parent_order_id': o.parent_order_id,
                'parent_order_number': o.parent_order.order_number if o.parent_order else '',
            })
        return JsonResponse({'orders': result})
    except Exception as e:
        return JsonResponse({'orders': [], 'error': str(e)})


@login_required(login_url='/staff/login/')
def live_pos_orders_api(request):
    """Returns ONLY active POS terminal cart snapshots (PosDraft building orders)
    for the staff portal live sidebar. Customer QR scan orders are NOT shown here
    - they appear via the sound notification popup instead.
    """
    try:
        import json as _json
        LIVE_MARKER = '__LIVE__'
        result = []

        # Only POS terminal drafts being actively built right now
        live_drafts = PosDraft.objects.filter(
            note__startswith=LIVE_MARKER,
            is_deleted=False,
        ).order_by('created_at')

        for d in live_drafts:
            try:
                items = _json.loads(d.items_json or '[]')
            except Exception:
                items = []
            if not items:
                continue   # skip empty carts
            items_preview = ', '.join(
                f"{i.get('qty',1)}× {i.get('name','?')}" for i in items[:3]
            )
            is_takeaway = (d.table_name or '').startswith('Takeaway')
            result.append({
                'id': f'draft-{d.id}',
                'draft_id': d.id,
                'order_number': d.draft_number,
                'table': d.table_name or 'POS',
                'table_name': d.table_name or 'POS',
                'order_type': 'takeaway' if is_takeaway else 'dine_in',
                'status': 'building',
                'total': float(d.total_amount),
                'items_preview': items_preview,
                'item_count': len(items),
            })

        return JsonResponse({'orders': result})
    except Exception as e:
        return JsonResponse({'orders': [], 'error': str(e)})


# --- EXCEL EXPORT -------------------------------------------------


@login_required(login_url='/staff/login/')
def pos_portal(request):
    """POS Table Selection page - entry point for POS at /pos/."""
    shop = get_shop()
    return render(request, 'restaurant/pos_table_selection.html', {
        'shop':        shop,
        'table_range': range(1, 16),   # Order 1-15
        'tw_range':    range(1, 6),    # Takeaway 1-5
    })


@login_required(login_url='/staff/login/')
def pos_terminal(request):
    """POS Terminal - full order entry UI at /pos/terminal/."""
    tables     = Table.objects.filter(is_active=True).order_by('number')
    categories = Category.objects.filter(is_active=True).prefetch_related('items')
    discounts  = Discount.objects.filter(is_active=True)
    combos     = Combo.objects.filter(is_active=True).prefetch_related('combo_items__menu_item')
    shop       = get_shop()

    # Build shop JSON for JS bill printing
    import json as _json
    shop_json = _json.dumps({
        'name':    shop.shop_name    if shop else 'Brothers Cafe',
        'address': shop.address      if shop else '',
        'phone':   shop.phone        if shop else '',
        'gstin':   shop.gstin        if shop else '',
        'fssai':   shop.fssai_number if shop else '',
        'upi':     shop.upi_id       if shop else '',
    })

    default_parcel = float(shop.default_parcel_charge) if shop and shop.default_parcel_charge else 0

    today       = timezone.now().date()
    week_start  = today - timezone.timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    def _rev(order_qs):
        # root.total_amount = session total after completion
        return sum(o.total_amount for o in order_qs)

    import datetime as _dt2
    _tz2 = timezone.get_current_timezone()
    _ts  = timezone.make_aware(_dt2.datetime.combine(today, _dt2.time.min), _tz2)
    _te  = timezone.make_aware(_dt2.datetime.combine(today, _dt2.time.max), _tz2)
    _ws  = timezone.make_aware(_dt2.datetime.combine(week_start, _dt2.time.min), _tz2)
    _ms  = timezone.make_aware(_dt2.datetime.combine(month_start, _dt2.time.min), _tz2)
    today_qs    = Order.objects.filter(status='completed', parent_order__isnull=True, created_at__range=(_ts, _te))
    week_qs     = Order.objects.filter(status='completed', parent_order__isnull=True, created_at__gte=_ws)
    month_qs    = Order.objects.filter(status='completed', parent_order__isnull=True, created_at__gte=_ms)
    all_qs      = Order.objects.filter(status='completed', parent_order__isnull=True)

    today_revenue   = _rev(today_qs)
    today_completed = today_qs.count()
    week_revenue    = _rev(week_qs)
    month_revenue   = _rev(month_qs)
    all_revenue     = _rev(all_qs)
    all_cnt         = all_qs.count()
    avg_order_val   = (all_revenue / all_cnt) if all_cnt else Decimal('0')

    top_items = (
        OrderItem.objects
        .filter(order__status='completed', order__parent_order__isnull=True, order__created_at__date__gte=month_start)
        .values(item_name=F_db('menu_item__name'))
        .annotate(
            total_qty=Sum('quantity'),
            total_rev=Sum(ExpressionWrapper(F_db('unit_price') * F_db('quantity'), output_field=DecimalField()))
        )
        .order_by('-total_qty')[:8]
    )

    from datetime import timedelta
    chart_labels, chart_values = [], []
    for i in range(6, -1, -1):
        d  = today - timedelta(days=i)
        qs = Order.objects.filter(status='completed', parent_order__isnull=True, created_at__date=d)
        chart_labels.append(d.strftime('%d %b'))
        chart_values.append(float(_rev(qs)))

    cart_offers_pos = active_cart_offers()
    offer_banners_pos = active_offer_banners()
    import json as _json_pos
    cart_offers_pos_json = _json_pos.dumps([{
        'id': co.id, 'title': co.title, 'min_cart_value': float(co.min_cart_value),
        'reward_type': co.reward_type, 'percent_off': float(co.percent_off),
        'flat_off': float(co.flat_off), 'emoji': co.emoji, 'reward_label': co.reward_label,
        'free_item_id': co.free_item_id, 'free_item_name': co.free_item.name if co.free_item else '',
        'free_item_price': float(co.free_item.price) if co.free_item else 0,
    } for co in cart_offers_pos])
    offer_banners_pos_json = _json_pos.dumps([{
        'id': b.id, 'title': b.title, 'offer_type': b.offer_type,
        'off_percent': float(b.off_percent), 'flat_amount': float(b.flat_amount),
        'offer_label': b.offer_label, 'emoji': b.emoji,
        'items': [{'id': mi.id, 'name': mi.name, 'price': float(mi.price),
                   'image': mi.image.url if mi.image else ''} for mi in b.menu_items.all()]
    } for b in offer_banners_pos])

    return render(request, 'restaurant/pos_terminal.html', {
        'tables':               tables,
        'categories':           categories,
        'discounts':            discounts,
        'combos':               combos,
        'shop':                 shop,
        'shop_json':            shop_json,
        'default_parcel':       default_parcel,
        'default_parcel_charge': default_parcel,
        'order_slots':          range(1, 16),
        'takeaway_slots':       range(1, 6),
        'cart_offers':          cart_offers_pos,
        'cart_offers_json':     cart_offers_pos_json,
        'offer_banners':        offer_banners_pos,
        'offer_banners_json':   offer_banners_pos_json,
        'today_revenue':        today_revenue,
        'today_completed':      today_completed,
        'week_revenue':         week_revenue,
        'month_revenue':        month_revenue,
        'all_revenue':          all_revenue,
        'avg_order_val':        avg_order_val,
        'top_items':            top_items,
        'chart_labels_json':    json.dumps(chart_labels),
        'chart_values_json':    json.dumps(chart_values),
    })

@login_required(login_url='/staff/login/')
def export_excel(request):
    try:
        orders = Order.objects.all().select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        wb = _build_excel(orders, 'All Orders')
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="brothers_cafe_all_orders.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)


@login_required(login_url='/staff/login/')
def export_daily_excel(request):
    try:
        import json as _json
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO as _BytesIO
        today = timezone.localdate()   # use local date (IST), not UTC
        orders = Order.objects.filter(created_at__date=today).select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        drafts = PosDraft.objects.filter(is_deleted=False, created_at__date=today).order_by('-created_at')

        wb = _build_excel(orders, 'Orders')

        # Sheet 2: POS Draft Orders
        ws2 = wb.create_sheet("POS Saved Orders")
        hdr_fill = PatternFill("solid", fgColor="1a1a2e")
        hdr_font = Font(color="FFFFFF", bold=True)
        headers = ['Draft #', 'Time', 'Customer', 'Phone', 'Table', 'Items', 'Subtotal', 'Discount%', 'Total', 'Note']
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for row, d in enumerate(drafts, 2):
            items_str = ', '.join([f"{i.get('qty',1)}x {i.get('name','?')}" for i in _json.loads(d.items_json)])
            ws2.append([d.draft_number, d.created_at.strftime('%H:%M'), d.customer_name,
                        d.customer_phone, d.table_name, items_str,
                        float(d.subtotal), float(d.discount_pct), float(d.total_amount), d.note])
        if drafts.exists():
            ws2.append([])
            ws2.append(['','','','','','TOTAL','',
                        '', float(sum(d.total_amount for d in drafts))])

        buf = _BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="brothers_cafe_{today}.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)






@login_required(login_url='/staff/login/')
def download_qr_codes(request):
    import io, re as _re

    base_url = request.build_absolute_uri('/').rstrip('/')

    tables_data = [
        {"label": t.name, "url": f"{base_url}/table/{t.id}/menu/", "takeaway": False}
        for t in Table.objects.filter(is_active=True).order_by('number')
    ]
    tables_data.append({"label": "Takeaway", "url": f"{base_url}/takeaway/", "takeaway": True})

    def make_qr_svg(url, size=210):
        """Generate QR code: tries segno (pure-python), then qrcode lib, then API image."""
        # 1. Try segno (pure Python, no C deps)
        try:
            import segno
            qr = segno.make_qr(url, error='M')
            buf = io.BytesIO()
            qr.save(buf, kind='svg', scale=4, border=2)
            svg = buf.getvalue().decode('utf-8')
            svg = _re.sub(r'<\?xml[^>]+\?>', '', svg)
            svg = _re.sub(r'<!DOCTYPE[^>]+>', '', svg)
            svg = _re.sub(r'width="[^"]+"', f'width="{size}"', svg, count=1)
            svg = _re.sub(r'height="[^"]+"', f'height="{size}"', svg, count=1)
            return svg.strip()
        except ImportError:
            pass
        except Exception:
            pass
        # 2. Try qrcode library
        try:
            import qrcode
            import qrcode.image.svg as qr_svg
            factory = qr_svg.SvgPathImage
            img = qrcode.make(url, image_factory=factory, box_size=10, border=2)
            buf = io.BytesIO()
            img.save(buf)
            svg = buf.getvalue().decode('utf-8')
            svg = _re.sub(r'<\?xml[^>]+\?>', '', svg)
            svg = _re.sub(r'<!DOCTYPE[^>]+>', '', svg)
            svg = _re.sub(r'width="[^"]+"', f'width="{size}px"', svg, count=1)
            svg = _re.sub(r'height="[^"]+"', f'height="{size}px"', svg, count=1)
            return svg.strip()
        except ImportError:
            pass
        except Exception:
            pass
        # 3. Fallback: use online QR API (works without any library)
        import urllib.parse as _up
        enc = _up.quote(url, safe='')
        return f'<img src="https://api.qrserver.com/v1/create-qr-code/?size={size}x{size}&data={enc}" width="{size}" height="{size}" style="display:block;border-radius:4px" alt="QR Code"/>'

    cards = []
    for t in tables_data:
        svg    = make_qr_svg(t["url"])
        color  = "#e67e22" if t["takeaway"] else "#1a1a2e"
        accent = "#fff8f0" if t["takeaway"] else "#f4f6ff"
        icon   = "🛍️" if t["takeaway"] else "🍽️"
        sub    = "Scan to order takeaway" if t["takeaway"] else "Scan to order food"
        cards.append(f"""<div class="qr-card">
<div class="card-top" style="background:{color}">
  <div class="card-icon">{icon}</div>
  <div><div class="card-name">Brothers Cafe</div><div class="card-sub">{sub}</div></div>
</div>
<div class="card-qr" style="background:{accent}">{svg}</div>
<div class="card-bottom" style="background:{color}">
  <span class="scan-text">coffee Scan &amp; Order</span>
  <span class="staff-label">Staff: {t['label']}</span>
</div></div>""")

    html = '''<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Brothers Cafe QR Codes</title>
<style>
@page{margin:15mm}
@media print{.no-print{display:none!important}body{background:white;padding:0}}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#eef0f3;padding:28px 20px}
h1{text-align:center;font-size:24px;font-weight:800;color:#1a1a2e;margin-bottom:6px}
.sub{text-align:center;font-size:13px;color:#666;margin-bottom:26px}
.grid{display:grid;grid-template-columns:repeat(3,1fr);gap:20px;max-width:820px;margin:0 auto}
.qr-card{background:white;border-radius:16px;overflow:hidden;box-shadow:0 4px 18px rgba(0,0,0,.13);break-inside:avoid}
.card-top{padding:14px 16px;display:flex;align-items:center;gap:10px;color:white}
.card-icon{font-size:26px}.card-name{font-size:14px;font-weight:800}.card-sub{font-size:11px;opacity:.75;margin-top:2px}
.card-qr{padding:16px;display:flex;align-items:center;justify-content:center}
.card-bottom{padding:10px 16px;color:white;display:flex;align-items:center;justify-content:space-between}
.scan-text{font-size:13px;font-weight:700}
.staff-label{font-size:10px;background:rgba(255,255,255,.18);padding:3px 9px;border-radius:20px}
.print-btn{display:block;margin:24px auto 0;padding:13px 38px;background:linear-gradient(135deg,#1a1a2e,#0f3460);color:white;border:none;border-radius:50px;font-size:15px;font-weight:700;cursor:pointer;font-family:inherit}
</style></head><body>
<h1>coffee Brothers Cafe - QR Codes</h1>
<p class="sub">Print &amp; place on each table. Customers scan to order - table numbers never shown to customers.</p>
<div class="grid">''' + ''.join(cards) + '''</div>
<button class="print-btn no-print" onclick="window.print()">🖨️ Print All QR Codes</button>
</body></html>'''

    return HttpResponse(html, content_type='text/html; charset=utf-8')

@login_required(login_url='/staff/login/')
def export_weekly_excel(request):
    try:
        today = timezone.localdate()   # use local date (IST), not UTC
        week_start = today - timezone.timedelta(days=today.weekday())
        orders = Order.objects.filter(created_at__date__gte=week_start).select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        wb = _build_excel(orders, f'Weekly {week_start} to {today}')
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="brothers_cafe_weekly_{week_start}.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)


@login_required(login_url='/staff/login/')
def export_monthly_excel(request):
    try:
        today = timezone.localdate()   # use local date (IST), not UTC
        month_start = today.replace(day=1)
        orders = Order.objects.filter(created_at__date__gte=month_start).select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        wb = _build_excel(orders, f'Monthly {today.strftime("%B %Y")}')
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="brothers_cafe_{today.strftime("%Y_%m")}.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)


@login_required(login_url='/staff/login/')
def export_range_excel(request):
    try:
        from datetime import datetime
        date_from_str = request.GET.get('from', '')
        date_to_str = request.GET.get('to', '')
        if not date_from_str or not date_to_str:
            return HttpResponse('Missing from/to parameters', status=400)
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        orders = Order.objects.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        ).select_related('table').prefetch_related('items__menu_item').order_by('-created_at')
        title = f'{date_from.strftime("%d %b")} to {date_to.strftime("%d %b %Y")}'
        wb = _build_excel(orders, title)
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="brothers_cafe_{date_from_str}_to_{date_to_str}.xlsx"'
        return resp
    except Exception as e:
        import traceback; traceback.print_exc()
        return HttpResponse(f'Export error: {e}', status=500)

def _build_excel(orders, title):
    """Build Excel. Each root-order session is ONE row (reorder children are
    merged in so they do not create duplicate rows)."""
    from openpyxl.styles import Font, Alignment, PatternFill

    # Separate roots from children so each session produces only one row.
    order_list = list(orders)
    child_ids = {o.id for o in order_list if o.parent_order_id is not None}
    root_orders = [o for o in order_list if o.parent_order_id is None]

    # Map child orders by their root id (walk up to handle multi-level nesting)
    child_by_root = {}
    for o in order_list:
        if o.parent_order_id is not None:
            root_id = o.parent_order_id
            # Walk up
            parent_map = {x.id: x for x in order_list}
            while root_id in parent_map and parent_map[root_id].parent_order_id:
                root_id = parent_map[root_id].parent_order_id
            child_by_root.setdefault(root_id, []).append(o)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Title
    ws.merge_cells('A1:Q1')
    ws['A1'] = f'Brothers Cafe - {title}'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='1a1a2e')
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Order #','Type','Date','Time','Customer','Phone','Table',
               'Items','Subtotal (Rs)','Offer / Discount','Disc (Rs)','Parcel (Rs)','Total (Rs)',
               'Status','Payment','Cash Received','Change Due']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='e74c3c')
        c.alignment = Alignment(horizontal='center')

    row = 3
    total_rev = Decimal('0')
    for order in root_orders:
        # Collect all items: root's own + every child's own (since items are no longer merged)
        all_chain_orders = [order] + child_by_root.get(order.id, [])
        order_items = []
        for co in all_chain_orders:
            order_items.extend(list(co.items.select_related('menu_item','combo').all()))
        items_str = ', '.join(f"{i.quantity}x {i.display_name}" for i in order_items)
        ldt = timezone.localtime(order.created_at)

        # Session total = each order in the chain minus its OWN discount
        def _order_total(o, items):
            sub = sum((i.unit_price * i.quantity for i in items), Decimal('0'))
            if o.discount_is_flat:
                disc = min(o.discount_amount or Decimal('0'), sub)
            elif o.discount_percent:
                disc = (sub * o.discount_percent / 100).quantize(Decimal('0.01'))
            else:
                disc = Decimal('0')
            return sub, disc, o.parcel_charge or Decimal('0')

        # Build per-order breakdown for the offer label
        offer_parts = []
        real_subtotal = Decimal('0')
        real_disc = Decimal('0')
        parcel = Decimal('0')
        for co in all_chain_orders:
            co_items = [i for i in order_items if i.order_id == co.id]
            sub, disc, par = _order_total(co, co_items)
            real_subtotal += sub
            real_disc += disc
            parcel += par
            if co.offer_title or co.discount_percent or disc > 0:
                lbl = co.offer_title or ''
                if co.discount_percent:
                    lbl += f' {int(co.discount_percent)}%'
                elif co.discount_is_flat:
                    lbl += f' flat Rs{int(disc)}'
                offer_parts.append(lbl.strip())
        offer_label = ' + '.join(p for p in offer_parts if p) or ''
        if not offer_label and real_disc > 0:
            offer_label = 'Discount'

        # real_disc and real_subtotal were already computed per-order in the loop above.
        # Don't reset real_disc here — the loop already applied each order's offer correctly.
        # Only recompute if the loop found no offers at all (legacy single-offer path).
        real_subtotal = sum((i.unit_price * i.quantity for i in order_items), Decimal('0'))
        if not offer_parts and real_disc == Decimal('0'):
            # No offer detected in the loop — fall back to reading from root order
            if order.discount_is_flat:
                real_disc = min(order.discount_amount or Decimal('0'), real_subtotal)
                offer_label = (order.offer_title or 'Flat off') + ' (flat Rs ' + str(int(real_disc)) + ')'
            elif order.discount_percent:
                real_disc = (real_subtotal * order.discount_percent / 100).quantize(Decimal('0.01'))
                offer_label = ((order.offer_title or '') + ' ' + str(int(order.discount_percent)) + '% off').strip()
        real_total = max(Decimal('0'), real_subtotal + parcel - real_disc)
        # Display combo names from display_name
        items_str = ', '.join(
            f"{i.quantity}x {i.display_name}" for i in order_items
        )

        ws.cell(row=row, column=1,  value=order.order_number)
        ws.cell(row=row, column=2,  value=order.get_order_type_display())
        ws.cell(row=row, column=3,  value=ldt.strftime('%d-%m-%Y'))
        ws.cell(row=row, column=4,  value=ldt.strftime('%H:%M'))
        ws.cell(row=row, column=5,  value=order.customer_name)
        ws.cell(row=row, column=6,  value=order.customer_phone)
        ws.cell(row=row, column=7,  value=str(order.table) if order.table else 'Takeaway')
        ws.cell(row=row, column=8,  value=items_str)
        ws.cell(row=row, column=9,  value=float(real_subtotal))
        ws.cell(row=row, column=10, value=offer_label)
        ws.cell(row=row, column=11, value=float(real_disc))
        ws.cell(row=row, column=12, value=float(parcel))
        ws.cell(row=row, column=13, value=float(real_total))
        ws.cell(row=row, column=14, value=order.get_status_display())
        ws.cell(row=row, column=15, value=order.get_payment_status_display())
        ws.cell(row=row, column=16, value=float(order.cash_received or 0))
        ws.cell(row=row, column=17, value=float(order.change_amount or 0))
        if order.status == 'completed':
            total_rev += real_total
        row += 1

    row += 1
    ws.cell(row=row, column=12, value='Total Revenue:').font = Font(bold=True)
    ws.cell(row=row, column=13, value=float(total_rev)).font = Font(bold=True)

    # Set column widths
    col_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q']
    for letter in col_letters:
        ws.column_dimensions[letter].width = 15
    ws.column_dimensions['H'].width = 48  # Items
    ws.column_dimensions['J'].width = 32  # Offer / Discount
    ws.column_dimensions['N'].width = 18  # Status
    return wb


# ===================================================================
# ITEMS PORTAL - Zomato/Swiggy-style menu management for admins
# ===================================================================

@login_required(login_url='/staff/login/')
def items_portal(request):
    """Admin-facing items management portal grouped by category."""
    categories = Category.objects.prefetch_related('items').order_by('order', 'name')
    shop = get_shop()
    return render(request, 'restaurant/items_portal.html', {
        'categories': categories,
        'shop': shop,
    })


@login_required(login_url='/staff/login/')
def items_portal_toggle(request, item_id):
    """Toggle dine-in / takeaway / featured for an item via AJAX."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        import json as _json
        data = _json.loads(request.body)
        field = data.get('field')
        item = MenuItem.objects.get(pk=item_id)
        if field == 'dine_in':
            item.is_available_dine_in = not item.is_available_dine_in
            item.save(update_fields=['is_available_dine_in'])
            return JsonResponse({'success': True, 'value': item.is_available_dine_in})
        elif field == 'takeaway':
            item.is_available_takeaway = not item.is_available_takeaway
            item.save(update_fields=['is_available_takeaway'])
            return JsonResponse({'success': True, 'value': item.is_available_takeaway})
        elif field == 'featured':
            item.is_featured = not item.is_featured
            item.save(update_fields=['is_featured'])
            return JsonResponse({'success': True, 'value': item.is_featured})
        return JsonResponse({'error': 'Unknown field'}, status=400)
    except MenuItem.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def items_portal_update_price(request, item_id):
    """Quick-update price for a menu item via AJAX."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        import json as _json
        data = _json.loads(request.body)
        price = Decimal(str(data.get('price', 0)))
        if price < 0:
            return JsonResponse({'error': 'Price must be non-negative'}, status=400)
        item = MenuItem.objects.get(pk=item_id)
        item.price = price
        item.save(update_fields=['price'])
        return JsonResponse({'success': True, 'price': float(price)})
    except MenuItem.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def items_portal_delete_item(request, item_id):
    """Delete a menu item via AJAX."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        item = MenuItem.objects.get(pk=item_id)
        item.delete()
        return JsonResponse({'success': True})
    except MenuItem.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ===================================================================
# ITEMS PORTAL - Add / Edit / Upload image - inline (no admin redirect)
# ===================================================================

@login_required(login_url='/staff/login/')
def items_portal_add_item(request):
    """Add a new menu item via AJAX (supports image upload)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        name         = request.POST.get('name', '').strip()
        category_id  = request.POST.get('category_id', '')
        price        = request.POST.get('price', '0')
        parcel_charge= request.POST.get('parcel_charge', '0')
        item_type    = request.POST.get('item_type', 'veg')
        description  = request.POST.get('description', '').strip()
        dine_in      = request.POST.get('is_available_dine_in', 'true') == 'true'
        takeaway     = request.POST.get('is_available_takeaway', 'true') == 'true'
        featured     = request.POST.get('is_featured', 'false') == 'true'

        if not name:
            return JsonResponse({'error': 'Item name is required'}, status=400)
        category = Category.objects.get(pk=category_id)

        image_file = request.FILES.get('image')

        item = MenuItem(
            name=name,
            category=category,
            price=Decimal(price),
            parcel_charge=Decimal(parcel_charge),
            item_type=item_type,
            description=description,
            is_available_dine_in=dine_in,
            is_available_takeaway=takeaway,
            is_featured=featured,
            is_water_bottle=request.POST.get('is_water_bottle', 'false') == 'true' or 'water bottle' in name.lower(),
            is_available=True,
        )
        # Save without image first so the item always exists in DB
        item.image = None
        item.save()

        image_url = None
        image_error = None
        if image_file:
            try:
                item.image = image_file
                item.save(update_fields=['image'])
                # Safely get URL - Cloudinary returns https when SECURE=True in settings
                try:
                    image_url = item.image.url if item.image else None
                except Exception:
                    image_url = None
            except Exception as img_err:
                import traceback; traceback.print_exc()
                image_error = str(img_err)
                item.image = None  # ensure field is clear

        response = {
            'success': True,
            'item_id': item.id,
            'name': item.name,
            'price': float(item.price),
            'image_url': image_url,
            'item_type': item.item_type,
        }
        if image_error:
            response['image_warning'] = f'Item saved but image upload failed: {image_error}'
        return JsonResponse(response)
    except Category.DoesNotExist:
        return JsonResponse({'error': 'Category not found'}, status=404)
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def items_portal_edit_item(request, item_id):
    """Edit a menu item via AJAX (supports image upload)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        item = MenuItem.objects.get(pk=item_id)

        name         = request.POST.get('name', item.name).strip()
        category_id  = request.POST.get('category_id', item.category_id)
        price        = request.POST.get('price', str(item.price))
        parcel_charge= request.POST.get('parcel_charge', str(item.parcel_charge))
        item_type    = request.POST.get('item_type', item.item_type)
        description  = request.POST.get('description', item.description).strip()
        dine_in      = request.POST.get('is_available_dine_in', 'true') == 'true'
        takeaway     = request.POST.get('is_available_takeaway', 'true') == 'true'
        featured     = request.POST.get('is_featured', 'false') == 'true'

        item.name          = name
        item.category_id   = category_id
        item.price         = Decimal(price)
        item.parcel_charge = Decimal(parcel_charge)
        item.item_type     = item_type
        item.description   = description
        item.is_available_dine_in   = dine_in
        item.is_available_takeaway  = takeaway
        item.is_featured            = featured
        item.is_water_bottle        = request.POST.get('is_water_bottle', 'false') == 'true'
        if not item.is_water_bottle and 'water bottle' in item.name.lower():
            item.is_water_bottle = True

        # Save item data first (no image) - always succeeds.
        # Then try saving the image separately so data is never lost.
        image_file = request.FILES.get('image')
        item.save()  # Saves name, price, category etc - always works

        image_url = None
        image_error = None
        if image_file:
            try:
                item.image = image_file
                item.save(update_fields=['image'])
                try:
                    image_url = item.image.url if item.image else None
                except Exception:
                    image_url = None
            except Exception as img_err:
                import traceback; traceback.print_exc()
                image_error = str(img_err)
        else:
            try:
                image_url = item.image.url if item.image else None
            except Exception:
                image_url = None

        response = {
            'success': True,
            'name': item.name,
            'price': float(item.price),
            'image_url': image_url,
        }
        if image_error:
            response['image_warning'] = f'Item saved but image upload failed: {image_error}'
        return JsonResponse(response)
    except MenuItem.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def items_portal_get_item(request, item_id):
    """Return item data as JSON for the edit modal."""
    try:
        item = MenuItem.objects.select_related('category').get(pk=item_id)
        try:
            image_url = item.image.url if item.image else None
        except Exception:
            image_url = None
        return JsonResponse({
            'id': item.id,
            'name': item.name,
            'category_id': item.category_id,
            'price': float(item.price),
            'parcel_charge': float(item.parcel_charge),
            'item_type': item.item_type,
            'description': item.description,
            'is_available_dine_in': item.is_available_dine_in,
            'is_available_takeaway': item.is_available_takeaway,
            'is_featured': item.is_featured,
            'is_water_bottle': item.is_water_bottle,
            'image_url': image_url,
        })
    except MenuItem.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)


# ===================================================================
# MENU MANAGER - Full inline CRUD APIs (no Django admin)
# ===================================================================

# -- CATEGORIES --------------------------------------------------
@login_required(login_url='/staff/login/')
def mm_categories(request):
    cats = list(Category.objects.values('id','name','icon','description','order','is_active'))
    for c in cats:
        c['item_count'] = MenuItem.objects.filter(category_id=c['id']).count()
    return JsonResponse({'categories': cats})

@login_required(login_url='/staff/login/')
def mm_category_save(request, cat_id=None):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        data = json.loads(request.body)
        if cat_id:
            cat = Category.objects.get(pk=cat_id)
        else:
            cat = Category()
        cat.name = data.get('name','').strip()
        cat.icon = data.get('icon','').strip()
        cat.description = data.get('description','').strip()
        cat.order = int(data.get('order', cat.order if cat_id else 0))
        cat.is_active = data.get('is_active', True)
        cat.save()
        return JsonResponse({'success':True,'id':cat.id,'name':cat.name})
    except Category.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

@login_required(login_url='/staff/login/')
def mm_category_toggle(request, cat_id):
    """Quick toggle is_active for a category without changing other fields."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        cat = Category.objects.get(pk=cat_id)
        cat.is_active = not cat.is_active
        cat.save(update_fields=['is_active'])
        return JsonResponse({'success': True, 'is_active': cat.is_active})
    except Category.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required(login_url='/staff/login/')
def mm_category_delete(request, cat_id):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        Category.objects.get(pk=cat_id).delete()
        return JsonResponse({'success':True})
    except Category.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

# -- COMBOS ------------------------------------------------------
@login_required(login_url='/staff/login/')
def mm_combos(request):
    combos = []
    for c in Combo.objects.prefetch_related('combo_items__menu_item').all():
        combos.append({
            'id':c.id,'name':c.name,'description':c.description,
            'price':float(c.price),'icon':c.icon,'is_active':c.is_active,'order':c.order,
            'is_offer':c.is_offer,'offer_tag':c.offer_tag,
            'image':c.image.url if c.image else '',
            'valid_from':c.valid_from.isoformat() if c.valid_from else '',
            'valid_to':c.valid_to.isoformat() if c.valid_to else '',
            'item_count':c.combo_items.count(),
            'items':[{'menu_item_id':ci.menu_item_id,'name':ci.menu_item.name,
                      'qty':ci.quantity,'quantity':ci.quantity,
                      'price':float(ci.menu_item.price)} for ci in c.combo_items.all()],
            'items_value':float(sum(ci.menu_item.price*ci.quantity for ci in c.combo_items.all()))
        })
    return JsonResponse({'combos':combos})

@login_required(login_url='/staff/login/')
def mm_combo_save(request, combo_id=None):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        # Accept FormData (so combos and combo offers can carry an image) as
        # well as plain JSON.
        ct = request.content_type or ''
        image_file = None
        remove_image = False
        if 'multipart' in ct:
            data = request.POST.dict()
            image_file = request.FILES.get('image')
            remove_image = str(data.get('remove_image','')).lower() in ('true','1')
            if data.get('items'):
                data['items'] = json.loads(data['items'])
            for _b in ('is_active','is_offer'):
                if _b in data:
                    data[_b] = str(data[_b]).lower() in ('true','1')
        else:
            data = json.loads(request.body)
        if combo_id:
            combo = Combo.objects.get(pk=combo_id)
        else:
            combo = Combo()
        combo.name = data.get('name','').strip()
        combo.description = data.get('description','').strip()
        combo.price = Decimal(str(data.get('price',0)))
        combo.icon = data.get('icon','🎁').strip()
        combo.is_active = data.get('is_active', True)
        combo.order = int(data.get('order', combo.order if combo_id else 0))
        combo.is_offer = bool(data.get('is_offer', combo.is_offer if combo_id else False))
        combo.offer_tag = (data.get('offer_tag') or '').strip()[:40]
        combo.valid_from = data.get('valid_from') or None
        combo.valid_to = data.get('valid_to') or None
        if combo.valid_from and combo.valid_to and combo.valid_to < combo.valid_from:
            return JsonResponse({'error': 'End date cannot be before the start date'}, status=400)
        if not combo.name:
            return JsonResponse({'error': 'Combo name is required'}, status=400)
        if combo.price is None or combo.price < 0:
            return JsonResponse({'error': 'Combo price must be zero or more'}, status=400)
        if remove_image and combo.image:
            combo.image.delete(save=False)
            combo.image = None
        if image_file:
            combo.image = image_file
        combo.save()

        # Items chosen by staff: [{'menu_item_id': 4, 'quantity': 2}, ...]
        # Previously there was no way to set these outside Django admin, so every
        # combo created from the manager had a price but no contents.
        items = data.get('items')
        if items is not None:
            wanted = {}
            for row in items:
                try:
                    mid = int(row.get('menu_item_id') or row.get('id'))
                    qty = max(1, int(row.get('quantity') or row.get('qty') or 1))
                except (TypeError, ValueError):
                    continue
                wanted[mid] = wanted.get(mid, 0) + qty
            valid_ids = set(MenuItem.objects.filter(id__in=wanted.keys())
                            .values_list('id', flat=True))
            combo.combo_items.exclude(menu_item_id__in=valid_ids).delete()
            for mid in valid_ids:
                ComboItem.objects.update_or_create(
                    combo=combo, menu_item_id=mid,
                    defaults={'quantity': wanted[mid]},
                )

        combo.refresh_from_db()
        return JsonResponse({
            'success': True,
            'id': combo.id,
            'item_count': combo.combo_items.count(),
            'items': [{'menu_item_id': ci.menu_item_id, 'name': ci.menu_item.name,
                       'quantity': ci.quantity, 'price': float(ci.menu_item.price)}
                      for ci in combo.combo_items.select_related('menu_item').all()],
            'items_value': float(combo.items_value),
            'savings': float(combo.savings),
            'is_offer': combo.is_offer,
            'offer_tag': combo.offer_tag,
            'valid_from': combo.valid_from.isoformat() if combo.valid_from else '',
            'valid_to': combo.valid_to.isoformat() if combo.valid_to else '',
            'image': combo.image.url if combo.image else '',
        })
    except Combo.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

@login_required(login_url='/staff/login/')
def mm_combo_toggle(request, combo_id):
    """Quick toggle is_active for a combo without changing other fields."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        combo = Combo.objects.get(pk=combo_id)
        combo.is_active = not combo.is_active
        combo.save(update_fields=['is_active'])
        return JsonResponse({'success': True, 'is_active': combo.is_active})
    except Combo.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required(login_url='/staff/login/')
def mm_combo_delete(request, combo_id):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        Combo.objects.get(pk=combo_id).delete()
        return JsonResponse({'success':True})
    except Combo.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

# -- DISCOUNTS ----------------------------------------------------
@login_required(login_url='/staff/login/')
def mm_discounts(request):
    discs = list(Discount.objects.values('id','name','percent','is_active','valid_from','valid_to','description'))
    for d in discs:
        d['valid_from'] = str(d['valid_from']) if d['valid_from'] else ''
        d['valid_to'] = str(d['valid_to']) if d['valid_to'] else ''
        d['percent'] = float(d['percent'])
    return JsonResponse({'discounts':discs})

@login_required(login_url='/staff/login/')
def mm_discount_save(request, disc_id=None):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        data = json.loads(request.body)
        if disc_id:
            disc = Discount.objects.get(pk=disc_id)
        else:
            disc = Discount()
        disc.name = data.get('name','').strip()
        disc.percent = Decimal(str(data.get('percent',0)))
        disc.is_active = data.get('is_active', True)
        disc.description = data.get('description','').strip()
        vf = data.get('valid_from','').strip()
        vt = data.get('valid_to','').strip()
        disc.valid_from = vf if vf else None
        disc.valid_to = vt if vt else None
        disc.save()
        return JsonResponse({'success':True,'id':disc.id})
    except Discount.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

@login_required(login_url='/staff/login/')
def mm_discount_delete(request, disc_id):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        Discount.objects.get(pk=disc_id).delete()
        return JsonResponse({'success':True})
    except Discount.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

# -- OFFER BANNERS ------------------------------------------------
@login_required(login_url='/staff/login/')
def mm_offers(request):
    offers = list(OfferBanner.objects.values('id','title','subtitle','offer_type','off_percent','flat_amount','emoji','bg_color','image_url','is_active','order'))
    for o in offers:
        o['off_percent'] = float(o['off_percent'])
        o['flat_amount'] = float(o['flat_amount'])
        o['menu_item_ids'] = list(OfferBanner.objects.get(pk=o['id']).menu_items.values_list('id', flat=True))
    return JsonResponse({'offers': offers})

@login_required(login_url='/staff/login/')
def mm_offer_save(request, offer_id=None):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        # Support both JSON and FormData (for image upload)
        ct = request.content_type or ''
        if 'multipart' in ct:
            data = request.POST.dict()
            image_file = request.FILES.get('image')
        else:
            data = json.loads(request.body)
            image_file = None

        if offer_id:
            ob = OfferBanner.objects.get(pk=offer_id)
        else:
            ob = OfferBanner()
        title = data.get('title','').strip()
        if not title:
            return JsonResponse({'error':'Title is required'},status=400)
        offer_type = data.get('offer_type', 'percent')
        if offer_type not in ('percent', 'flat', 'bogo'):
            offer_type = 'percent'
        if offer_type == 'percent':
            pct = Decimal(str(data.get('off_percent', 0) or 0))
            if pct < 0 or pct > 100:
                return JsonResponse({'error':'OFF % must be 0-100'},status=400)
            ob.off_percent = pct
            ob.flat_amount = Decimal('0')
        elif offer_type == 'flat':
            flat = Decimal(str(data.get('flat_amount', 0) or 0))
            if flat < 0:
                return JsonResponse({'error':'Flat amount must be positive'},status=400)
            ob.flat_amount = flat
            ob.off_percent = Decimal('0')
        else:
            ob.off_percent = Decimal('0')
            ob.flat_amount = Decimal('0')
        ob.title = title
        ob.subtitle = data.get('subtitle','').strip()
        ob.offer_type = offer_type
        ob.emoji = (data.get('emoji','') or '🎉').strip() or '🎉'
        ob.bg_color = (data.get('bg_color','') or '#e74c3c').strip() or '#e74c3c'
        ob.image_url = (data.get('image_url','') or '').strip()
        ob.is_active = str(data.get('is_active', 'true')).lower() in ('true','1','on','yes')
        ob.valid_from = data.get('valid_from') or None
        ob.valid_to = data.get('valid_to') or None
        if ob.valid_from and ob.valid_to and str(ob.valid_to) < str(ob.valid_from):
            return JsonResponse({'error': 'End date cannot be before the start date'}, status=400)
        try:
            ob.order = int(data.get('order', 0) or 0)
        except Exception:
            ob.order = 0
        # Save first (so we have a PK for M2M)
        ob.save()
        # Handle image upload
        image_url_resp = ob.banner_image_url
        if image_file:
            try:
                ob.image = image_file
                ob.save(update_fields=['image'])
                image_url_resp = ob.banner_image_url
            except Exception as img_err:
                import traceback; traceback.print_exc()
        # Remove image if requested
        if data.get('remove_image') == 'true' and not image_file:
            ob.image = None
            ob.image_url = ''
            ob.save(update_fields=['image', 'image_url'])
            image_url_resp = ''
        # Link menu items
        item_ids_raw = data.get('menu_item_ids', '[]')
        if isinstance(item_ids_raw, str):
            try:
                item_ids = json.loads(item_ids_raw)
            except Exception:
                item_ids = []
        else:
            item_ids = item_ids_raw
        if isinstance(item_ids, list):
            ob.menu_items.set(MenuItem.objects.filter(id__in=item_ids))
        return JsonResponse({'success':True,'id':ob.id,'image_url':image_url_resp})
    except OfferBanner.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'error':str(e)},status=500)

@login_required(login_url='/staff/login/')
def mm_offer_toggle(request, offer_id):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        ob = OfferBanner.objects.get(pk=offer_id)
        ob.is_active = not ob.is_active
        ob.save()
        return JsonResponse({'success':True,'is_active':ob.is_active})
    except OfferBanner.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)


@ensure_csrf_cookie
def offer_page(request, offer_id):
    """Dedicated offer page - shows offer details and linked items for ordering."""
    offer = get_object_or_404(OfferBanner, id=offer_id, is_active=True)
    items = offer.menu_items.filter(is_available=True).select_related('category')
    shop = get_shop()
    # Get table/session context
    table_id = request.GET.get('table')
    reorder_id = request.GET.get('reorder')
    table = None
    if table_id:
        table = Table.objects.filter(id=table_id, is_active=True).first()
    return render(request, 'restaurant/offer_page.html', {
        'offer': offer,
        'items': items,
        'table': table,
        'reorder_id': reorder_id or '',
        'shop': shop,
    })

@login_required(login_url='/staff/login/')
def mm_offer_delete(request, offer_id):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        OfferBanner.objects.get(pk=offer_id).delete()
        return JsonResponse({'success':True})
    except OfferBanner.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)


# -- CART OFFERS (conditional cart-level promotions) --------------
@login_required(login_url='/staff/login/')
def mm_cart_offers(request):
    offers = list(CartOffer.objects.values('id','title','subtitle','min_cart_value','reward_type','percent_off','flat_off','free_item_id','emoji','is_active','order'))
    for o in offers:
        o['min_cart_value'] = float(o['min_cart_value'])
        o['percent_off'] = float(o['percent_off'])
        o['flat_off'] = float(o['flat_off'])
        # get free item name
        if o['free_item_id']:
            try: o['free_item_name'] = MenuItem.objects.get(pk=o['free_item_id']).name
            except MenuItem.DoesNotExist: o['free_item_name'] = ''
        else: o['free_item_name'] = ''
    return JsonResponse({'cart_offers': offers})


@login_required(login_url='/staff/login/')
def mm_cart_offer_save(request, cart_offer_id=None):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        data = json.loads(request.body)
        if cart_offer_id:
            co = CartOffer.objects.get(pk=cart_offer_id)
        else:
            co = CartOffer()
        title = data.get('title','').strip()
        if not title: return JsonResponse({'error':'Title is required'},status=400)
        co.title = title
        co.subtitle = data.get('subtitle','').strip()
        co.min_cart_value = Decimal(str(data.get('min_cart_value',0) or 0))
        rt = data.get('reward_type','percent')
        if rt not in ('percent','flat','free_item'): rt = 'percent'
        co.reward_type = rt
        co.percent_off = Decimal(str(data.get('percent_off',0) or 0))
        co.flat_off = Decimal(str(data.get('flat_off',0) or 0))
        fi_id = data.get('free_item_id')
        co.free_item = MenuItem.objects.get(pk=fi_id) if fi_id else None
        co.emoji = (data.get('emoji','') or '🎁').strip() or '🎁'
        co.is_active = str(data.get('is_active','true')).lower() in ('true','1')
        try: co.order = int(data.get('order',0) or 0)
        except Exception: co.order = 0
        co.valid_from = data.get('valid_from') or None
        co.valid_to = data.get('valid_to') or None
        if co.valid_from and co.valid_to and str(co.valid_to) < str(co.valid_from):
            return JsonResponse({'error': 'End date cannot be before the start date'}, status=400)
        co.save()
        # Dates arrive as strings from JSON; re-read so the response returns
        # real date objects rather than blowing up on .isoformat().
        co.refresh_from_db()
        return JsonResponse({'success':True,'id':co.id,
                             'valid_from': co.valid_from.isoformat() if co.valid_from else '',
                             'valid_to': co.valid_to.isoformat() if co.valid_to else ''})
    except CartOffer.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)


@login_required(login_url='/staff/login/')
def mm_cart_offer_delete(request, cart_offer_id):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        CartOffer.objects.get(pk=cart_offer_id).delete()
        return JsonResponse({'success':True})
    except CartOffer.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)


# -- TABLES -------------------------------------------------------
@login_required(login_url='/staff/login/')
def mm_tables(request):
    tables = list(Table.objects.values('id','number','name','capacity','status','description','is_active'))
    return JsonResponse({'tables':tables})

@login_required(login_url='/staff/login/')
def mm_table_save(request, table_id=None):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        data = json.loads(request.body)
        if table_id:
            tbl = Table.objects.get(pk=table_id)
        else:
            tbl = Table()
        tbl.number = int(data.get('number', tbl.number if table_id else 1))
        tbl.name = data.get('name','').strip()
        tbl.capacity = int(data.get('capacity', 4))
        tbl.description = data.get('description','').strip()
        tbl.is_active = data.get('is_active', True)
        tbl.save()
        return JsonResponse({'success':True,'id':tbl.id})
    except Table.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

@login_required(login_url='/staff/login/')
def mm_table_delete(request, table_id):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        Table.objects.get(pk=table_id).delete()
        return JsonResponse({'success':True})
    except Table.DoesNotExist: return JsonResponse({'error':'Not found'},status=404)
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

# -- CUSTOMERS ----------------------------------------------------
@login_required(login_url='/staff/login/')
def mm_customers(request):
    custs = []
    for c in CustomerProfile.objects.all()[:100]:
        custs.append({
            'id':c.id,'name':c.name,'phone':c.phone,
            'visit_count':c.visit_count,
            'total_orders':c.total_orders,
            'total_spent':float(c.total_spent),
            'last_visit':c.last_visit.strftime('%d %b %Y'),
        })
    return JsonResponse({'customers':custs})

# -- SHOP SETTINGS ------------------------------------------------
@login_required(login_url='/staff/login/')
def mm_shop_settings(request):
    shop = ShopSettings.objects.first()
    if not shop: return JsonResponse({'settings':None})
    return JsonResponse({'settings':{
        'id':shop.id,'shop_name':shop.shop_name,'location':shop.location,
        'gstin':shop.gstin,'fssai_number':shop.fssai_number,
        'phone':shop.phone,'email':shop.email,'address':shop.address,
        'upi_id':shop.upi_id,
        'default_discount_percent':float(shop.default_discount_percent),
        'default_parcel_charge':float(shop.default_parcel_charge),
    }})

@login_required(login_url='/staff/login/')
def mm_shop_settings_save(request):
    if request.method != 'POST': return JsonResponse({'error':'POST required'},status=405)
    try:
        # Accept FormData (needed for the logo upload) as well as plain JSON.
        logo_file = None
        if 'multipart' in (request.content_type or ''):
            data = request.POST.dict()
            logo_file = request.FILES.get('logo')
            for _b in ('show_water_bottle_in_cart',):
                if _b in data:
                    data[_b] = str(data[_b]).lower() in ('true','1','on','yes')
        else:
            data = json.loads(request.body)
        shop = ShopSettings.objects.first()
        if not shop: shop = ShopSettings()
        shop.shop_name = data.get('shop_name','Brothers Cafe').strip()
        shop.location = data.get('location','').strip()
        shop.gstin = data.get('gstin','').strip()
        shop.fssai_number = data.get('fssai_number','').strip()
        shop.phone = data.get('phone','').strip()
        shop.email = data.get('email','').strip()
        shop.address = data.get('address','').strip()
        shop.upi_id = data.get('upi_id','').strip()
        shop.default_discount_percent = Decimal(str(data.get('default_discount_percent',0)))
        shop.default_parcel_charge = Decimal(str(data.get('default_parcel_charge',0)))

        # -- Water bottle quick-add card ----------------------------------
        # These were never saved, so the toggle and price in the manager had no
        # effect and the card fell back to whatever beverage it could find.
        if 'show_water_bottle_in_cart' in data:
            shop.show_water_bottle_in_cart = str(data.get('show_water_bottle_in_cart')).lower() in ('true','1','on','yes')
        if 'water_bottle_item_id' in data:
            wb_id = data.get('water_bottle_item_id')
            shop.water_bottle_item = MenuItem.objects.filter(pk=wb_id, is_water_bottle=True).first() if wb_id else None
        # A pinned item that is no longer flagged must not keep being used.
        if shop.water_bottle_item and not shop.water_bottle_item.is_water_bottle:
            shop.water_bottle_item = None
        if 'water_bottle_cart_price' in data:
            try:
                price = Decimal(str(data.get('water_bottle_cart_price') or 0))
            except Exception:
                price = Decimal('0')
            if price < 0:
                return JsonResponse({'error': 'Water bottle price cannot be negative'}, status=400)
            shop.water_bottle_cart_price = price
        if shop.show_water_bottle_in_cart and not shop.water_bottle_item \
                and not MenuItem.objects.filter(is_water_bottle=True).exists():
            return JsonResponse({'error': 'No water bottle available. Add a menu item and tick '
                                          '"This is a water bottle" on it first.'}, status=400)


        if logo_file:
            shop.logo = logo_file
        shop.save()
        return JsonResponse({'success': True,
                             'logo': shop.logo.url if shop.logo else ''})
    except Exception as e: return JsonResponse({'error':str(e)},status=500)

# -- ORDERS LIST --------------------------------------------------
@login_required(login_url='/staff/login/')
def mm_orders(request):
    page = int(request.GET.get('page',1))
    status = request.GET.get('status','')
    qs = Order.objects.select_related('table').order_by('-created_at')
    if status: qs = qs.filter(status=status)
    from django.core.paginator import Paginator
    p = Paginator(qs, 20)
    pg = p.get_page(page)
    orders = []
    for o in pg:
        orders.append({
            'id':o.id,'order_number':o.order_number,'customer_name':o.customer_name,
            'customer_phone':o.customer_phone,'table':str(o.table) if o.table else 'Takeaway',
            'status':o.status,'order_type':o.order_type,
            'total_amount':float(o.total_amount),
            'created_at':o.created_at.strftime('%d %b %Y %H:%M'),
            'payment_status':o.payment_status,
        })
    return JsonResponse({'orders':orders,'total_pages':p.num_pages,'page':page})

# -- MENU MANAGER MAIN PAGE ----------------------------------------
@login_required(login_url='/staff/login/')
@login_required(login_url='/staff/login/')
def mm_category_reorder(request):
    """Save new drag-and-drop order for categories."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        data = json.loads(request.body)
        ids = data.get('ids', [])  # ordered list of category IDs
        for idx, cat_id in enumerate(ids):
            Category.objects.filter(pk=cat_id).update(order=idx + 1)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def mm_items_by_category(request, cat_id):
    """Return menu items for a category, ordered by current order field."""
    try:
        cat = Category.objects.get(pk=cat_id)
        items = list(
            MenuItem.objects.filter(category=cat).order_by('order', 'name')
            .values('id', 'name', 'price', 'item_type', 'is_available_dine_in', 'is_available_takeaway', 'order')
        )
        return JsonResponse({'items': items, 'category': {'id': cat.id, 'name': cat.name, 'icon': cat.icon}})
    except Category.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def mm_items_reorder(request):
    """Save new drag-and-drop order for menu items."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        data = json.loads(request.body)
        ids = data.get('ids', [])  # ordered list of menu item IDs
        for idx, item_id in enumerate(ids):
            MenuItem.objects.filter(pk=item_id).update(order=idx + 1)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/staff/login/')
def menu_manager(request):
    """Main menu manager page - renders the full management UI."""
    categories = Category.objects.prefetch_related('items').order_by('order','name')
    menu_items = MenuItem.objects.select_related('category').order_by('category__name','name')
    all_categories = Category.objects.order_by('order','name')
    all_combos = Combo.objects.prefetch_related('combo_items__menu_item').all()
    combos = [c for c in all_combos if not c.is_offer]
    combo_offers = [c for c in all_combos if c.is_offer]
    discounts = Discount.objects.all()
    offers = OfferBanner.objects.prefetch_related('menu_items').all()
    cart_offers = CartOffer.objects.all()
    tables = Table.objects.order_by('number')
    customers = CustomerProfile.objects.all()[:50]
    shop = ShopSettings.objects.first()

    # Stats
    total_items = MenuItem.objects.count()
    available_items = MenuItem.objects.filter(is_available_dine_in=True).count()
    avg_price = MenuItem.objects.aggregate(a=models_Avg('price'))['a'] or 0

    picker_items = [
        {'id': m.id, 'name': m.name, 'price': float(m.price),
         'category': m.category.name if m.category else 'Other',
         'type': m.item_type or 'veg'}
        for m in MenuItem.objects.select_related('category').order_by('category__name', 'name')
    ]

    # Items flagged as water bottle, OR (fallback) items whose name contains
    # "water bottle" — so the card works even if the toggle wasn't ticked yet.
    water_candidates = MenuItem.objects.filter(is_water_bottle=True).select_related('category').order_by('name')
    if not water_candidates.exists():
        water_candidates = MenuItem.objects.filter(name__icontains='water bottle').select_related('category').order_by('name')

    return render(request, 'restaurant/menu_manager.html', {
        'water_candidates': water_candidates,
        'categories': categories,
        'menu_items': menu_items,
        'all_categories': all_categories,
        'combos': combos,
        'picker_items': picker_items,
        'combo_offers': combo_offers,
        'combo_count': len(combos),
        'combo_offer_count': len(combo_offers),
        'discounts': discounts,
        'offers': offers,
        'cart_offers': cart_offers,
        'tables': tables,
        'customers': customers,
        'shop': shop,
        'total_items': total_items,
        'available_items': available_items,
        'avg_price': avg_price,
        'cat_count': all_categories.count(),
    })
